import streamlit as st
import pandas as pd
import datetime
import sys
import os
import json
import io
import time
import locale

from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode
from calendar import monthrange 

# --- Adicionar o diretório raiz ao path ---
try:
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    diretorio_pai = os.path.dirname(diretorio_atual)
    if diretorio_pai not in sys.path:
        sys.path.append(diretorio_pai)

    import calculos as calc
    import processamento_dados as proc_dados
    import graficos as gfx 
except ImportError:
    st.error("Erro fatal: Não foi possível encontrar os módulos locais (calculos.py, processamento_dados.py). Certifique-se de que este ficheiro está na pasta 'pages' e os outros estão no diretório principal.")
    st.stop()


# --- Configuração da Página ---
st.set_page_config(page_title="Simulador de Tarifários Gás Natural 2025: Poupe na Fatura | Tiago Felícia", page_icon="🔥", layout="wide", initial_sidebar_state="collapsed")

# --- Carregar Dados ---
url_excel = "https://github.com/tiagofelicia/simulador-tarifarios-gas/raw/refs/heads/main/Tarifarios_%F0%9F%94%A5_Gas_Natural_Tiago_Felicia.xlsx"

# Configura a ordenação para seguir as regras do Português
try:
    locale.setlocale(locale.LC_COLLATE, 'pt_PT.UTF-8')
except locale.Error:
    st.warning("Aviso: O 'locale' Português não está instalado no sistema. A ordenação de municípios com acentos pode não estar correta.")

try:
    (
        CONSTANTES, tarifas_gas_master, tos_municipios, mibgas_df, info_tab
    ) = proc_dados.carregar_dados_excel_gas(url_excel)
    
    if tarifas_gas_master.empty or CONSTANTES.empty or tos_municipios.empty:
        st.error("Erro: Uma das abas essenciais ('Tarifas_Gas_Master', 'Constantes', 'TOS') não foi carregada ou está vazia.")
        st.stop()
    if mibgas_df.empty:
        st.warning("Aviso: A aba 'MIBGAS' não foi carregada. Os tarifários indexados a gás podem não ser calculados.")

except ValueError:
    st.error("Erro fatal ao carregar os dados.")
    st.stop()
except Exception as e:
    st.error(f"Ocorreu um erro ao carregar os dados do Excel: {e}")
    st.stop()

# --- Obter valor constante da Quota ACP ---
VALOR_QUOTA_ACP_MENSAL = calc.obter_constante("Quota_ACP", CONSTANTES)

#FUNÇÕES
# --- Mapas para encurtar os parâmetros do URL ---
MAPA_ESCALAO_PARA_URL = {
    "Escalão 1 (Consumo até 220 m³/ano)": "E1",
    "Escalão 2 (Consumo 221 a 500 m³/ano)": "E2",
    "Escalão 3 (Consumo 501 a 1.000 m³/ano)": "E3",
    "Escalão 4 (Consumo 1.001 a 10.000 m³/ano)": "E4",
}
# Cria o dicionário inverso automaticamente para ler os URLs
MAPA_URL_PARA_ESCALAO = {v: k for k, v in MAPA_ESCALAO_PARA_URL.items()}

def inicializar_estado_e_url_gas():
    """
    Verifica e inicializa o st.session_state para o simulador de Gás, dando prioridade a valores no URL. Corre apenas uma vez por sessão.
    """
    if 'estado_inicializado_gas' in st.session_state:
        return

    # 1. Obter lista de municípios para validação
    lista_municipios_validos = sorted(tos_municipios['Município'].dropna().unique())
    
    # 2. Ler Escalão do URL, ou usar default
    escalao_codigo_url = st.query_params.get("esc") # Pega o código, ex: "E4"
    nome_longo_escalao = MAPA_URL_PARA_ESCALAO.get(escalao_codigo_url) # Procura "E4" no dicionário e devolve o nome completo

    # Se encontrou um nome de escalão correspondente no dicionário, usa-o.
    if nome_longo_escalao:
        st.session_state.sel_escalao_gas_key = nome_longo_escalao
    else:
        # Caso contrário (parâmetro inválido ou ausente), usa o default.
        st.session_state.sel_escalao_gas_key = "Escalão 1 (Consumo até 220 m³/ano)"

    # 3. Ler Município do URL, ou usar default
    municipio_url = st.query_params.get("mun")
    if municipio_url and municipio_url in lista_municipios_validos:
        st.session_state.sel_municipio_tos = municipio_url
    else:
        # Usar Almeirim (10º da lista - índice 9) como default, com fallback
        default_idx = 9 if len(lista_municipios_validos) > 9 else 0
        st.session_state.sel_municipio_tos = lista_municipios_validos[default_idx]

    # 4. Ler Consumo e Modo de Input
    st.session_state.gas_input_mode = "Consumo (m³)" if "con_m3" in st.query_params else "Consumo (kWh)"
    st.session_state.gas_kwh_input_key = float(st.query_params.get("con_kwh", 135)) # Default 135 kWh

    # 5. Ler Preço MIBGAS
    if "mibgas" in st.query_params:
        st.session_state.mibgas_input_mwh_manual = float(st.query_params.get("mibgas"))
        
    # 6. Ler Opções Adicionais
    st.session_state.chk_ts_gas_v2 = st.query_params.get("ts") == "1"
    st.session_state.chk_acp_gas = st.query_params.get("acp") != "0" # Default é True
    st.session_state.chk_cont_gas = st.query_params.get("cont") != "0" # Default é True

    # 7. Ler Tarifário Personalizado ---
    if st.query_params.get("p_a") == "1":
        st.session_state.chk_pers_gas_ativo = True
        st.session_state.pers_gas_energia = float(st.query_params.get("p_e", 0.0))
        st.session_state.pers_gas_fixo = float(st.query_params.get("p_f", 0.0))
        st.session_state.pers_gas_tar_energia = st.query_params.get("p_te") != "0" # Default é True
        st.session_state.pers_gas_tar_potencia = st.query_params.get("p_tf") != "0" # Default é True

    # Adicionar uma flag para indicar que a inicialização foi concluída
    st.session_state.estado_inicializado_gas = True

# --- Chamada da Função de Inicialização ---
inicializar_estado_e_url_gas()

# --- Função para REINICIAR o simulador para os valores padrão ---
def reiniciar_simulador():
    """
    Repõe o simulador de GÁS para os valores padrão, limpando o session_state e os parâmetros do URL.
    """
    
    # Lista de todas as chaves de session_state específicas do simulador de Gás
    chaves_a_apagar_gas = [
        # Chave de inicialização
        'estado_inicializado_gas',
        
        # Chaves de Inputs Principais
        'sel_escalao_gas_key',
        'sel_municipio_tos',
        'gas_input_mode',
        'gas_kwh_input_key',
        'gas_m3_input_key',
        'gas_pcs_input_key',
        
        # Chaves de Datas e MIBGAS
        'sel_mes_gas',
        'session_initialized_dates_gas',
        'data_inicio_key_input_gas',
        'data_fim_key_input_gas',
        'dias_manual_input_key_gas',
        'previous_mes_for_dates_gas',

        
        # Chaves MIBGAS
        'mibgas_input_mwh_manual',
        'mibgas_default_calculado',
        
        # Chaves de Opções Adicionais
        'gas_isp_manual_input',
        'chk_ts_gas_v2',
        'chk_acp_gas',
        'chk_cont_gas',
        
        # Chaves do "Meu Tarifário"
        'chk_meu_tarifario_gas_ativo',
        'meu_termo_energia_gas',
        'meu_termo_fixo_gas',
        'meu_gas_tar_energia_incluida',
        'meu_gas_tar_fixo_incluida',
        'meu_gas_desconto_energia_perc',
        'meu_gas_desconto_fixo_perc',
        'meu_gas_desconto_fatura_eur',
        'meu_gas_acrescimo_fatura_eur',
        
        # Chaves do "Tarifário Personalizado"
        'chk_pers_gas_ativo',
        'pers_gas_energia',
        'pers_gas_fixo',
        'pers_gas_tar_energia',
        'pers_gas_tar_potencia',
        
        # Chaves de Filtros e UI
        'filter_segmento_gas_idx',
        'filter_faturacao_gas_idx',
        'filter_pagamento_gas_idx',
        'filter_tipos_gas_multi',
        'chk_vista_simplificada_gas',
        'poupanca_excel_texto_gas',
        'poupanca_excel_cor_gas',
        'poupanca_excel_negrito_gas'
    ]

    # Apagar cada chave se ela existir
    for key in chaves_a_apagar_gas:
        if key in st.session_state:
            del st.session_state[key]

    # Limpar todos os parâmetros do URL
    st.query_params.clear()

    # Definir explicitamente o estado das checkboxes como Falso
    st.session_state.chk_meu_tarifario_gas_ativo = False
    st.session_state.chk_pers_gas_ativo = False

    st.success("Simulador de Gás reiniciado para os valores padrão.")

# --- Título e Botão de Limpeza Geral ---

# Linha 1: Logo e Título
col_logo, col_titulo = st.columns([1, 5])

with col_logo:
    st.image("https://raw.githubusercontent.com/tiagofelicia/simulador-tarifarios-eletricidade/refs/heads/main/Logo_Tiago_Felicia.png", width=180)

with col_titulo:
    st.title("🔥 Tiago Felícia - Simulador de Tarifários de Gás Natural")

st.button(
    "🧹 Limpar e Reiniciar Simulador",
    on_click=reiniciar_simulador,
    help="Repõe os campos do simulador para os valores iniciais.",
    use_container_width=True
)

# --- FUNÇÕES AUXILIARES PARA EXPORTAÇÃO EXCEL ---
def gerar_estilo_completo_para_valor(valor, minimo, maximo):
    """Gera o CSS completo (cor de fundo e texto) para uma célula de gradiente."""
    estilo_css_final = 'text-align: center;' 
    if pd.isna(valor): return estilo_css_final
    try: val_float = float(valor)
    except ValueError: return estilo_css_final
    if maximo == minimo or minimo is None or maximo is None: return estilo_css_final

    midpoint = (minimo + maximo) / 2
    r_bg, g_bg, b_bg = 255,255,255 
    verde_rgb, branco_rgb, vermelho_rgb = (99,190,123), (255,255,255), (248,105,107)

    if val_float <= midpoint:
        ratio = (val_float - minimo) / (midpoint - minimo) if midpoint != minimo else 0.0
        r_bg = int(verde_rgb[0]*(1-ratio) + branco_rgb[0]*ratio)
        g_bg = int(verde_rgb[1]*(1-ratio) + branco_rgb[1]*ratio)
        b_bg = int(verde_rgb[2]*(1-ratio) + branco_rgb[2]*ratio)
    else:
        ratio = (val_float - midpoint) / (maximo - midpoint) if maximo != midpoint else 0.0
        r_bg = int(branco_rgb[0]*(1-ratio) + vermelho_rgb[0]*ratio)
        g_bg = int(branco_rgb[1]*(1-ratio) + vermelho_rgb[1]*ratio)
        b_bg = int(branco_rgb[2]*(1-ratio) + vermelho_rgb[2]*ratio)

    estilo_css_final += f' background-color: #{r_bg:02X}{g_bg:02X}{b_bg:02X};'
    luminancia = (0.299 * r_bg + 0.587 * g_bg + 0.114 * b_bg)
    cor_texto_css = '#000000' if luminancia > 140 else '#FFFFFF'
    estilo_css_final += f' color: {cor_texto_css};'
    return estilo_css_final

def estilo_geral_dataframe_para_exportar(df_a_aplicar_estilo, tipos_reais_para_estilo_serie, min_max_config_para_cores, nome_coluna_tarifario="Tarifário"):
    """Aplica estilos CSS a todo o DataFrame pandas.Styler."""
    df_com_estilos = pd.DataFrame('', index=df_a_aplicar_estilo.index, columns=df_a_aplicar_estilo.columns)
    
    # Cores (Gás)
    cor_fundo_indexado_gas_css = "#FFE699"
    cor_texto_indexado_gas_css = "black"
    cor_fundo_fixo_gas_css = "#f0f0f0"
    cor_texto_fixo_gas_css = "#333333"

    for nome_coluna_df in df_a_aplicar_estilo.columns:
        if nome_coluna_df in min_max_config_para_cores:
            try:
                serie_valores_col = pd.to_numeric(df_a_aplicar_estilo[nome_coluna_df], errors='coerce')
                min_valor_col = min_max_config_para_cores[nome_coluna_df]['min']
                max_valor_col = min_max_config_para_cores[nome_coluna_df]['max']
                df_com_estilos[nome_coluna_df] = serie_valores_col.apply(
                    lambda valor_v: gerar_estilo_completo_para_valor(valor_v, min_valor_col, max_valor_col)
                )
            except Exception as e_estilo_custo:
                print(f"Erro ao aplicar estilo de custo à coluna {nome_coluna_df}: {e_estilo_custo}")
                df_com_estilos[nome_coluna_df] = 'text-align: center;' 

        elif nome_coluna_df == nome_coluna_tarifario: 
            estilos_col_tarif_lista = []
            for idx_linha_df, valor_nome_col_tarif in df_a_aplicar_estilo[nome_coluna_df].items():
                tipo_tarif_str = tipos_reais_para_estilo_serie.get(idx_linha_df, '') if tipos_reais_para_estilo_serie is not None else ''
                est_css_tarif = 'text-align: center; padding: 4px;' 
                bg_cor_val, fonte_cor_val, fonte_peso_val = "#FFFFFF", "#000000", "normal"

                if tipo_tarif_str == "Pessoal":
                    bg_cor_val, fonte_cor_val, fonte_peso_val = "#FF0000", "#FFFFFF", "bold"
                elif tipo_tarif_str == 'Indexado':
                    bg_cor_val, fonte_cor_val = cor_fundo_indexado_gas_css, cor_texto_indexado_gas_css
                elif tipo_tarif_str == 'Fixo': 
                    bg_cor_val, fonte_cor_val = cor_fundo_fixo_gas_css, cor_texto_fixo_gas_css
        
                est_css_tarif += f' background-color: {bg_cor_val}; color: {fonte_cor_val}; font-weight: {fonte_peso_val};'
                estilos_col_tarif_lista.append(est_css_tarif)
            df_com_estilos[nome_coluna_df] = estilos_col_tarif_lista
        else: 
            df_com_estilos[nome_coluna_df] = 'text-align: center;'
    return df_com_estilos

def exportar_excel_completo(df_para_exportar, styler_obj, resumo_html_para_excel, poupanca_texto_para_excel, identificador_cor_cabecalho, meu_tarifario_ativo_flag, personalizado_gas_ativo_flag):
    """Função Mestra de Exportação Excel"""
    output_excel_buffer = io.BytesIO() 
    with pd.ExcelWriter(output_excel_buffer, engine='openpyxl') as writer_excel:
        sheet_name_excel = 'Tiago Felicia - Gás Natural'

        # --- Escrever Resumo (Usa BeautifulSoup para ler o HTML do resumo) ---
        dados_resumo_formatado = []
        if resumo_html_para_excel:
            soup_resumo = BeautifulSoup(resumo_html_para_excel, "html.parser")
            titulo_resumo = soup_resumo.find('h5')
            if titulo_resumo:
                dados_resumo_formatado.append([titulo_resumo.get_text(strip=True), None])

            itens_lista_resumo = soup_resumo.find_all('li')
            linha_filtros_texto = ""
            linha_escalao_texto = ""
            outras_linhas_resumo = []

            for item in itens_lista_resumo:
                texto_item = item.get_text(separator=' ', strip=True)
                if "Segmento:" in texto_item:
                    linha_filtros_texto = texto_item
                elif "Escalão" in texto_item or "Município" in texto_item:
                    linha_escalao_texto = texto_item
                else:
                    parts = texto_item.split(':', 1)
                    if len(parts) == 2:
                        outras_linhas_resumo.append([parts[0].strip() + ":", parts[1].strip()])
                    else:
                        outras_linhas_resumo.append([texto_item, None])
            
            if linha_filtros_texto or linha_escalao_texto:
                dados_resumo_formatado.append([linha_filtros_texto, linha_escalao_texto])
            dados_resumo_formatado.extend(outras_linhas_resumo)
        
        df_resumo_obj = pd.DataFrame(dados_resumo_formatado)
        df_resumo_obj.to_excel(writer_excel, sheet_name=sheet_name_excel, index=False, header=False, startrow=0)
        worksheet_excel = writer_excel.sheets[sheet_name_excel]

        bold_font_obj = Font(bold=True)
        for i_resumo in range(len(df_resumo_obj)):
            excel_row_idx_resumo = i_resumo + 1
            cell_resumo_rotulo = worksheet_excel.cell(row=excel_row_idx_resumo, column=1)
            cell_resumo_rotulo.font = bold_font_obj
            if df_resumo_obj.shape[1] > 1 and pd.notna(df_resumo_obj.iloc[i_resumo, 1]):
                cell_resumo_valor = worksheet_excel.cell(row=excel_row_idx_resumo, column=2)
                cell_resumo_valor.font = bold_font_obj

        worksheet_excel.column_dimensions['A'].width = 35
        worksheet_excel.column_dimensions['B'].width = 65
        linha_atual_no_excel_escrita = len(df_resumo_obj) + 1

        # --- Escrever Mensagem de Poupança ---
        if poupanca_texto_para_excel:
            linha_atual_no_excel_escrita += 1 
            cor_p = st.session_state.get('poupanca_excel_cor_gas', "000000")
            negrito_p = st.session_state.get('poupanca_excel_negrito_gas', False)
            poupanca_cell_escrita = worksheet_excel.cell(row=linha_atual_no_excel_escrita, column=1, value=poupanca_texto_para_excel)
            poupanca_cell_escrita.font = Font(bold=negrito_p, color=cor_p)
            worksheet_excel.merge_cells(start_row=linha_atual_no_excel_escrita, start_column=1, end_row=linha_atual_no_excel_escrita, end_column=6)
            poupanca_cell_escrita.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
            linha_atual_no_excel_escrita += 1

        # --- Linha de Informação da Simulação ---
        linha_info_simulacao_excel = linha_atual_no_excel_escrita + 1 
        data_hoje_obj = datetime.date.today()
        data_hoje_formatada_str = data_hoje_obj.strftime('%d/%m/%Y')
        espacador_info = " " * 70
        texto_completo_info = f"          Simulação em {data_hoje_formatada_str}{espacador_info}https://www.tiagofelicia.pt{espacador_info}Tiago Felícia"
        info_cell = worksheet_excel.cell(row=linha_info_simulacao_excel, column=1)
        info_cell.value = texto_completo_info
        info_cell.font = Font(bold=True)
        worksheet_excel.merge_cells(start_row=linha_info_simulacao_excel, start_column=1, end_row=linha_info_simulacao_excel, end_column=6)
        info_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) 
        linha_inicio_tab_dados_excel = linha_info_simulacao_excel + 2 

        # Limpar fundo branco default
        for row in worksheet_excel.iter_rows(min_row=1, max_row=worksheet_excel.max_row+100, min_col=1, max_col=20):
            for cell in row:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Escrever o DataFrame Estilizado
        styler_obj.to_excel(
            writer_excel,
            sheet_name=sheet_name_excel,
            index=False,
            startrow=linha_inicio_tab_dados_excel - 1, 
            columns=df_para_exportar.columns.tolist()
        )

        # Cor do Cabeçalho (identificador_cor_cabecalho é o 'Escalão' para Gás)
        cor_fundo = "A6A6A6"; cor_fonte = "000000" # Default

        for col_idx, _ in enumerate(df_para_exportar.columns):
            celula = worksheet_excel.cell(row=linha_inicio_tab_dados_excel, column=col_idx + 1)
            celula.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
            celula.font = Font(color=cor_fonte, bold=True)

        # --- LEGENDA DE CORES ---
        ultima_linha_tabela_principal = linha_inicio_tab_dados_excel + len(df_para_exportar)
        linha_legenda_bloco_inicio = ultima_linha_tabela_principal + 2 

        titulo_legenda_cell = worksheet_excel.cell(row=linha_legenda_bloco_inicio, column=1, value="Tipos de Tarifário:")
        titulo_legenda_cell.font = Font(bold=True)
        worksheet_excel.merge_cells(start_row=linha_legenda_bloco_inicio, start_column=1, end_row=linha_legenda_bloco_inicio, end_column=6)
        titulo_legenda_cell.alignment = Alignment(horizontal='center', vertical='center')
        linha_legenda_item_atual = linha_legenda_bloco_inicio + 1 

        itens_legenda_excel = []
        # 1. Adicionar "O Meu Tarifário" se estiver ativo
        if meu_tarifario_ativo_flag:
            itens_legenda_excel.append(
                {"cf": "FF0000", "ct": "FFFFFF", "b": True, "tA": "O Meu Tarifário", "tB": "Tarifário configurado pelo utilizador."}
            )
        
        # 2. Adicionar "Tarifário Personalizado" se estiver ativo
        if personalizado_gas_ativo_flag:
             itens_legenda_excel.append(
                 {"cf": "92D050", "ct": "FFFFFF", "b": True, "tA": "Tarifário Personalizado", "tB": "Tarifário configurado pelo utilizador."}
             )

        # 3. Adicionar os tarifários base que aparecem sempre        
        itens_legenda_excel.extend([
            {"cf": "FFE699", "ct": "000000", "b": False, "tA": "Indexado", "tB": "Preço de energia baseado no MIBGAS + Margem."},
            {"cf": "F0F0F0", "ct": "333333", "b": False, "tA": "Fixo", "tB": "Preços de energia constantes", "borda_cor": "CCCCCC"}
        ])

        worksheet_excel.column_dimensions[get_column_letter(1)].width = 30
        worksheet_excel.column_dimensions[get_column_letter(2)].width = 200

        for item in itens_legenda_excel:
            celula_A_legenda = worksheet_excel.cell(row=linha_legenda_item_atual, column=1, value=item["tA"])
            celula_A_legenda.fill = PatternFill(start_color=item["cf"], end_color=item["cf"], fill_type="solid")
            celula_A_legenda.font = Font(color=item["ct"], bold=item["b"])
            celula_A_legenda.alignment = Alignment(horizontal='center', vertical='center', indent=1)
            if "borda_cor" in item:
                celula_A_legenda.border = Border(top=Side(style="thin", color=item["borda_cor"]), left=Side(style="thin", color=item["borda_cor"]), right=Side(style="thin", color=item["borda_cor"]), bottom=Side(style="thin", color=item["borda_cor"]))
            celula_B_legenda = worksheet_excel.cell(row=linha_legenda_item_atual, column=2, value=item["tB"])
            celula_B_legenda.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left')
            worksheet_excel.merge_cells(start_row=linha_legenda_item_atual, start_column=2, end_row=linha_legenda_item_atual, end_column=6) 
            worksheet_excel.row_dimensions[linha_legenda_item_atual].height = 20
            linha_legenda_item_atual += 1
        # --- FIM LEGENDA ---

        # Ajustar largura das colunas da tabela principal (Adaptado para Gás)
        for col_idx_iter, col_nome_iter_width in enumerate(df_para_exportar.columns):
            col_letra_iter = get_column_letter(col_idx_iter + 1) 
            if "Tarifário" in col_nome_iter_width:
                 worksheet_excel.column_dimensions[col_letra_iter].width = 80    
            elif "Total (€)" == col_nome_iter_width:
                worksheet_excel.column_dimensions[col_letra_iter].width = 25
            elif "(€/kWh)" in col_nome_iter_width or "(€/dia)" in col_nome_iter_width:
                worksheet_excel.column_dimensions[col_letra_iter].width = 25
            elif "Comercializador" in col_nome_iter_width:
                 worksheet_excel.column_dimensions[col_letra_iter].width = 30    
            elif "Faturação" in col_nome_iter_width:
                 worksheet_excel.column_dimensions[col_letra_iter].width = 33    
            elif "Pagamento" in col_nome_iter_width:
                 worksheet_excel.column_dimensions[col_letra_iter].width = 50    
            else: 
                worksheet_excel.column_dimensions[col_letra_iter].width = 25

    output_excel_buffer.seek(0)
    return output_excel_buffer


# --- Funções de Callback ---
def atualizar_consumo_default_gas():
    """
    Callback para pré-preencher o consumo anual baseado no escalão selecionado
    E ATUALIZAR O URL com o código do escalão.
    """
    escalao_str = st.session_state.get('sel_escalao_gas_key', "Escalão 1") 
    
    # Lógica de pré-preenchimento
    consumo_defaults = { 1: 135, 2: 300, 3: 600, 4: 1000 }
    escalao_num = escalao_map.get(escalao_str, 1)
    consumo_default = consumo_defaults.get(escalao_num, 0)
    st.session_state.gas_kwh_input_key = consumo_default

    # Lógica de atualização do URL
    codigo_escalao = MAPA_ESCALAO_PARA_URL.get(escalao_str)
    if codigo_escalao:
        st.query_params["esc"] = codigo_escalao

def atualizar_url_datas_gas():
    """Callback para monitorizar e atualizar o URL com as datas da simulação."""
    # Como as datas são complexas e interligadas, esta função não faz nada por agora.
    # A simples seleção já guarda o estado. Futuramente, poderíamos adicionar a lógica.
    pass

def sincronizar_datas_pelo_mes():
    """
    Callback executada quando o seletor de MÊS é alterado. Atualiza as datas de início e fim para corresponderem ao mês selecionado.
    """
    # Obter o mês selecionado a partir do estado da sessão
    mes_selecionado = st.session_state.sel_mes_gas
    
    # Obter o ano atual e o número do mês
    ano_atual = datetime.datetime.now().year
    mes_num = list(dias_mes.keys()).index(mes_selecionado) + 1
    
    # Calcular o primeiro e último dia do mês selecionado
    primeiro_dia = datetime.date(ano_atual, mes_num, 1)
    ultimo_dia = datetime.date(ano_atual, mes_num, dias_mes[mes_selecionado])
    
    # Atualizar diretamente as chaves dos widgets de data
    # Isto força os widgets a usarem as novas datas na próxima renderização
    st.session_state.data_inicio_key_input_gas = primeiro_dia
    st.session_state.data_fim_key_input_gas = ultimo_dia
    
    # Limpar o input manual de dias, pois as datas foram alteradas
    if 'dias_manual_input_key_gas' in st.session_state:
        del st.session_state['dias_manual_input_key_gas']

def sincronizar_mes_pelas_datas():
    """
    Callback executada quando um dos campos de DATA é alterado. Atualiza o seletor de mês para refletir a nova data de início.
    """
    # Obter a nova data de início a partir do estado da sessão do seu widget
    nova_data_inicio = st.session_state.data_inicio_key_input_gas
    
    # Obter a lista de meses
    meses_lista = list(dias_mes.keys())
    
    # Sincronizar o seletor de mês com o mês da nova data de início
    st.session_state.sel_mes_gas = meses_lista[nova_data_inicio.month - 1]
    
    # Limpar o input manual de dias
    if 'dias_manual_input_key_gas' in st.session_state:
        del st.session_state['dias_manual_input_key_gas']

def atualizar_url_mibgas():
    """Callback para o preço MIBGAS."""
    mibgas_default = st.session_state.get('mibgas_default_calculado', 30.0)
    mibgas_atual = st.session_state.get('mibgas_input_mwh_manual', mibgas_default)
    if mibgas_atual != mibgas_default:
        st.query_params['mibgas'] = str(mibgas_atual)
    elif 'mibgas' in st.query_params:
        del st.query_params['mibgas']

def atualizar_url_municipio():
    """Callback para o município selecionado."""
    municipio_selecionado = st.session_state.get('sel_municipio_tos')
    if municipio_selecionado:
        # Não verificamos default aqui, pois o município é um input principal
        st.query_params['mun'] = municipio_selecionado

def atualizar_url_consumo_gas():
    """Callback para os inputs de consumo (kWh ou m³)."""
    modo = st.session_state.get('gas_input_mode')
    if modo == "Consumo (kWh)":
        if 'con_m3' in st.query_params: del st.query_params['con_m3']
        if 'pcs' in st.query_params: del st.query_params['pcs']
        
        consumo_kwh = st.session_state.get('gas_kwh_input_key', 135)
        if consumo_kwh != 135: # Default
            st.query_params['con_kwh'] = str(consumo_kwh)
        elif 'con_kwh' in st.query_params:
            del st.query_params['con_kwh']

    elif modo == "Consumo (m³)":
        if 'con_kwh' in st.query_params: del st.query_params['con_kwh']
        
        consumo_m3 = st.session_state.get('gas_m3_input_key', 12)
        pcs = st.session_state.get('gas_pcs_input_key', 11.25)
        if consumo_m3 != 12: # Default
            st.query_params['con_m3'] = str(consumo_m3)
        elif 'con_m3' in st.query_params:
            del st.query_params['con_m3']
        
        if pcs != 11.25: # Default
            st.query_params['pcs'] = str(pcs)
        elif 'pcs' in st.query_params:
            del st.query_params['pcs']

def atualizar_url_opcoes_adicionais_gas():
    """Callback para todas as opções no expander de Opções Adicionais."""
    # Tarifa Social
    if st.session_state.get("chk_ts_gas_v2", False):
        st.query_params["ts"] = "1"
    elif "ts" in st.query_params:
        del st.query_params["ts"]
    
    # ACP e Continente (Default é True, guardamos no URL se for False)
    if not st.session_state.get("chk_acp_gas", True):
        st.query_params["acp"] = "0"
    elif "acp" in st.query_params:
        del st.query_params["acp"]

    if not st.session_state.get("chk_cont_gas", True):
        st.query_params["cont"] = "0"
    elif "cont" in st.query_params:
        del st.query_params["cont"]

def atualizar_url_meu_tarifario_gas():
    """Callback para os inputs do Meu Tarifário."""
    chaves_meu_tar = ["m_a", "m_e", "m_f", "m_te", "m_tf", "m_de", "m_dtf","m_df", "m_af"]
    for chave in chaves_meu_tar:
        if chave in st.query_params:
            del st.query_params[chave]

    if not st.session_state.get("chk_meu_tarifario_gas_ativo", False):
        return

    st.query_params["m_a"] = "1" # 'm_a' = meu_ativo

    # Preços
    if st.session_state.get("meu_termo_energia_gas"): st.query_params['m_e'] = st.session_state.get("meu_termo_energia_gas")
    if st.session_state.get("meu_termo_fixo_gas"): st.query_params['m_f'] = st.session_state.get("meu_termo_fixo_gas")
    # Flags (default=True)
    if not st.session_state.get("meu_gas_tar_energia_incluida", True): st.query_params['m_te'] = "0"
    if not st.session_state.get("meu_gas_tar_fixo_incluida", True): st.query_params['m_tf'] = "0"
    # Descontos/Acréscimos
    if st.session_state.get("meu_gas_desconto_energia_perc"): st.query_params['m_de'] = st.session_state.get("meu_gas_desconto_energia_perc")
    if st.session_state.get("meu_gas_desconto_fixo_perc"): st.query_params['m_dtf'] = st.session_state.get("meu_gas_desconto_fixo_perc")
    if st.session_state.get("meu_gas_desconto_fatura_eur"): st.query_params['m_df'] = st.session_state.get("meu_gas_desconto_fatura_eur")
    if st.session_state.get("meu_gas_acrescimo_fatura_eur"): st.query_params['m_af'] = st.session_state.get("meu_gas_acrescimo_fatura_eur")

def atualizar_url_tarifario_personalizado_gas():
    """Callback para os inputs do Tarifário Personalizado."""
    
    # 1. Lista de todas as chaves possíveis para este widget no URL
    chaves_pers_tar = ["p_a", "p_e", "p_f", "p_te", "p_tf"]
    
    # 2. Limpar sempre as chaves antigas do URL
    for chave in chaves_pers_tar:
        if chave in st.query_params:
            del st.query_params[chave]

    # 3. Se a secção não estiver ativa, sair
    if not st.session_state.get("chk_pers_gas_ativo", False):
        return

    # 4. Se estiver ativa, adicionar a flag 'p_a' e os outros parâmetros
    st.query_params["p_a"] = "1" # 'p_a' = personalizado_gas_ativo

    # Adicionar os preços apenas se forem > 0
    preco_energia = st.session_state.get("pers_gas_energia", 0.0)
    if preco_energia: 
        st.query_params['p_e'] = preco_energia

    preco_fixo = st.session_state.get("pers_gas_fixo", 0.0)
    if preco_fixo:
        st.query_params['p_f'] = preco_fixo
        
    # Adicionar as flags apenas se forem diferentes do default (que é True)
    if not st.session_state.get("pers_gas_tar_energia", True):
        st.query_params['p_te'] = "0"
        
    if not st.session_state.get("pers_gas_tar_potencia", True):
        st.query_params['p_tf'] = "0"

# ##################################################################
# INÍCIO DO BLOCO - GUIA RÁPIDO E FAQ
# ##################################################################

with st.expander("❓ Como Usar o Simulador de Tarifários de Gás Natural (Guia Rápido)", expanded=False):
    st.markdown("""
    Bem-vindo! Esta ferramenta ajuda-o a descobrir o tarifário de gás natural mais económico para si. Siga os passos abaixo para começar a poupar.

    #### **Passo 1: Defina o Período e o seu Perfil de Consumo**
    Primeiro, configure as bases da sua simulação.
    
    1.  **Selecione o Período:** Escolha o **mês** ou as **datas** para as quais pretende simular a sua fatura.
    2.  **Defina o Escalão e Município:** Indique o seu **Escalão de Consumo** (pode encontrá-lo na sua fatura) e o seu **Município**. O município é essencial para calcular corretamente a Taxa de Ocupação do Subsolo (TOS).
    3.  **Insira o Consumo:** Pode inserir o seu consumo de duas formas:
        * **Consumo (kWh):** O valor final de energia que aparece na sua fatura.
        * **Consumo (m³):** O volume de gás consumido, que também encontra na fatura. Terá de indicar o **Fator de Conversão (PCS)**, que converte m³ para kWh.

    #### **Passo 2: ⚙️ Refine a Simulação (Opcional)**
    Depois de inserir os seus dados principais, pode ajustar os detalhes.

    * **Preço MIBGAS:** Se estiver a analisar tarifários indexados, pode ajustar o preço médio esperado do MIBGAS (€/MWh). O simulador já sugere um valor com base em dados históricos e de futuros.
    * **Opções Adicionais:** No *expander* de "Opções Adicionais", pode ativar benefícios como a **Tarifa Social** (apenas para escalões 1 e 2) ou incluir descontos específicos de parcerias (ACP, Continente).
    * **O Meu Tarifário:** Use esta secção para introduzir os preços da sua fatura atual. Assim, pode compará-la diretamente com todas as ofertas do mercado e ver exatamente quanto pode poupar.

    #### **Passo 3: 🏆 Encontre a Melhor Tarifa**
    A tabela de resultados no final da página é a sua ferramenta principal.

    * **Ordenar por Custo:** Clique no cabeçalho da coluna **"Total Período (€)"** para ordenar os tarifários do mais barato para o mais caro.
    * **Explorar Detalhes:** Passe o rato sobre os preços (**Termo Fixo** e **Termo Energia**) ou sobre o **custo total** para ver um resumo detalhado dos cálculos, incluindo todas as taxas e impostos.
    * **Filtrar Resultados:** Use os filtros no topo da tabela para refinar a sua pesquisa por tipo de tarifário (Fixo, Indexado), segmento, etc.
    * **O Seu Pódio:** No final, a secção **"🏆 O Seu Pódio da Poupança"** destaca as 3 opções mais económicas para si.

    > **Dica Pro:**

    * **Use "O Meu Tarifário" como Ponto de Partida:** A forma mais poderosa de usar o simulador é introduzir os dados da sua fatura atual na secção "O Meu Tarifário". Verá imediatamente uma comparação direta e saberá se o seu contrato atual é competitivo ou exatamente quanto pode poupar ao mudar.
    * **Teste Cenários com o MIBGAS:** Se está a considerar um tarifário indexado, não altere o preço MIBGAS na primeira simulação. Depois, experimente inserir um valor mais alto (cenário pessimista) e um mais baixo (cenário otimista) para perceber a sensibilidade do custo final às variações do mercado.
    * **Não Ignore o Município:** A escolha do seu município é crucial. Afeta diretamente o valor da Taxa de Ocupação do Subsolo (TOS) e qual o Comercializador de Último Recurso (CUR) aplicável, influenciando o custo final da Tarifa Regulada.
    * **Explore os Filtros:** Se procura algo específico, como um tarifário com fatura eletrónica e débito direto, use os filtros no topo da tabela. Muitas vezes, as ofertas mais económicas encontram-se aqui."""      
    )

with st.expander("❔ Perguntas Frequentes (FAQ)", expanded=False):
    st.markdown("""
    ### Perguntas Gerais

    **P: De onde vêm os dados dos tarifários e dos preços de mercado (MIBGAS)?**
    
    **R:** Todos os dados dos tarifários são recolhidos a partir das informações públicas disponibilizadas pelos comercializadores nos seus websites. Os preços do mercado ibérico de gás (MIBGAS) são obtidos de fontes oficiais. Os dados são atualizados regularmente para refletir as condições atuais do mercado.

    **P: O simulador é 100% preciso?**
    
    **R:** O objetivo é ser o mais preciso possível. Para tarifários **fixos**, a precisão é muito elevada. Para tarifários **indexados**, o custo final é uma estimativa baseada nos preços médios do MIBGAS para o período selecionado. Variações diárias no preço do gás podem levar a pequenas diferenças no custo final.
    
    ---
    
    ### Dados e Períodos   
    **P: Qual a diferença entre escolher um Mês e um Período de Datas manual?**
    
    **R:** Estas são as diferenças:
    * **Escolher um Mês**: É a forma mais simples. O simulador seleciona automaticamente o primeiro e o último dia desse mês.
    * **Selecionar Datas**: Oferece total flexibilidade para analisar um período específico (ex: uma semana, uma quinzena).
                
    ---
    
    ### Funcionalidades do Simulador

    **P: Como funciona a secção "O Meu Tarifário"?**
    
    **R:** Esta secção permite-lhe introduzir os preços da sua fatura atual (**Termo Fixo** em €/dia e **Termo de Energia** em €/kWh) para comparar diretamente com todas as outras ofertas do mercado. É fundamental verificar na sua fatura se os preços que insere já incluem as **TAR (Tarifas de Acesso às Redes)**.

    **P: Onde encontro os valores do "Termo Fixo" e "Termo de Energia" na minha fatura?**

    **R:** Estes valores estão normalmente na secção de detalhe da sua fatura. Procure por:
    * **Termo Fixo (ou Termo de Disponibilidade):** Um valor em **€/dia**. Multiplique-o pelo número de dias do período de faturação para obter o custo total do termo fixo.
    * **Termo de Energia (ou Custo da Energia):** Um valor em **€/kWh**. Este é o preço que paga por cada kWh de gás consumido.
    Certifique-se de que insere os valores **sem IVA** para que o simulador possa fazer os cálculos corretamente.

    **P: Para que serve a secção "Comparar outro Tarifário Personalizado?"**
    
    **R:** Esta funcionalidade permite-lhe criar um cenário hipotético com os seus próprios preços. É ideal para simular uma oferta que recebeu e que ainda não está na lista ou para testar o impacto de futuras alterações de preços.
                
    **P: Os valores MIBGAS para datas futuras são reais?**
    
    **R:** Não. Os valores MIBGAS apresentados para datas futuras baseiam-se nos preços do mercado de futuros. Estes valores representam a expectativa do mercado, mas não são uma garantia. Servem como a melhor estimativa disponível para simular custos em tarifários indexados para períodos que ainda não ocorreram.

    **P: Com que frequência são atualizados os tarifários e os preços de mercado?**

    **R:** Os dados são atualizados regularmente para refletir as alterações nos preçários dos comercializadores e as variações do mercado de gás (MIBGAS). Pode consultar a data da última atualização dos valores de mercado na secção **"Datas de Referência"**, no final da página.                    

    **P: O que está incluído no "Total Período (€)"? É o valor final da fatura?**
    
    **R:** Sim, o valor "Total Período (€)" representa a sua fatura final estimada. Ele inclui a soma de várias componentes:
    1.  **Custo da Energia**: O consumo (kWh) multiplicado pelo Termo de Energia (€/kWh).
    2.  **Custo do Termo Fixo**: O Termo Fixo (€/dia) multiplicado pelo número de dias.
    3.  **Taxas e Impostos**: A Taxa de Ocupação do Subsolo (TOS), o Imposto sobre os Produtos Petrolíferos (ISP) e o IVA (aplicado a 6% ou 23% sobre cada componente, de acordo com as regras em vigor).
    
    Pode ver a decomposição detalhada de todos estes custos passando o rato por cima do valor na coluna "Total Período (€)".
                
    ---
    
    ### Termos e Conceitos

    **P: O que é um tarifário de gás indexado?**
    
    **R:** Um tarifário indexado tem um preço de energia que varia de acordo com o preço do mercado grossista de gás (MIBGAS). Este tipo de tarifário pode oferecer uma poupança significativa quando os preços de mercado estão baixos, mas também implica um risco maior se os preços subirem.

    **P: O que é o Escalão de Consumo?**
    
    **R:** É uma faixa de consumo anual (medido em m³) que determina os preços aplicados ao seu contrato. Existem 4 escalões para clientes domésticos e pequenos negócios. Pode encontrar o seu escalão na sua fatura de gás.

    **P: O que são as TAR (Tarifas de Acesso às Redes) no gás?**
    
    **R:** Tal como na eletricidade, as TAR no gás pagam pelo uso das infraestruturas (armazenamento, transporte e distribuição). São definidas pela ERSE. Alguns comercializadores apresentam os preços já com as TAR incluídas, enquanto outros as mostram em separado. O simulador lida com ambas as situações.

    **P: O que é o Fator de Conversão (PCS)?**
    
    **R:** O seu contador mede o gás em volume (metros cúbicos, m³), mas a energia que ele contém (o seu Poder Calorífico Superior, ou PCS) pode variar ligeiramente. O PCS é o fator que converte o volume (m³) em energia (kWh). Este valor está sempre indicado na sua fatura e é necessário se optar por inserir o seu consumo em m³.

    **P: O que é a TOS (Taxa de Ocupação do Subsolo)?**

    **R:** É uma taxa municipal pelo uso do subsolo público para a passagem das condutas de gás. O valor varia de município para município, sendo por isso crucial selecionar o seu corretamente.
    """)

# ##################################################################
# FIM DO BLOCO
# ##################################################################

# --- Inputs do Utilizador ---
st.subheader("⚙️ Insira os seus dados de simulação de Gás Natural")

# --- BLOCO DE DEFINIÇÃO DE DATAS ---
dias_mes = {"Janeiro":31,"Fevereiro":29,"Março":31,"Abril":30,"Maio":31,"Junho":30,"Julho":31,"Agosto":31,"Setembro":30,"Outubro":31,"Novembro":30,"Dezembro":31}
ano_atual_para_bissexto = datetime.datetime.now().year
if not ((ano_atual_para_bissexto % 4 == 0 and ano_atual_para_bissexto % 100 != 0) or (ano_atual_para_bissexto % 400 == 0)):
    dias_mes["Fevereiro"] = 28

# Lógica de inicialização (executada apenas uma vez)
if 'session_initialized_dates_gas' not in st.session_state:
    hoje = datetime.date.today()
    data_inicial_default = hoje + datetime.timedelta(days=1)
    
    ano_final, mes_final = (data_inicial_default.year, data_inicial_default.month + 1)
    if mes_final > 12:
        mes_final = 1
        ano_final += 1
    
    dias_no_mes_final = monthrange(ano_final, mes_final)[1]
    dia_final = min(data_inicial_default.day, dias_no_mes_final)
    
    data_final_bruta = datetime.date(ano_final, mes_final, dia_final)
    data_final_default = data_final_bruta - datetime.timedelta(days=1)
    
    # Define os valores iniciais para as chaves dos widgets
    st.session_state.sel_mes_gas = list(dias_mes.keys())[data_inicial_default.month - 1]
    st.session_state.data_inicio_key_input_gas = data_inicial_default
    st.session_state.data_fim_key_input_gas = data_final_default
    st.session_state.session_initialized_dates_gas = True

# --- Widgets com as novas callbacks ---
col_mes, col_data_i, col_data_f, col_dias_calc, col_dias_man = st.columns(5)

with col_mes:
    mes = st.selectbox(
        "Mês", 
        list(dias_mes.keys()), 
        key="sel_mes_gas", 
        on_change=sincronizar_datas_pelo_mes,
        help="Selecione um mês para preencher automaticamente as datas. Se o mês escolhido já tiver terminado, o valor do MIBGAS é final, se ainda estiver em curso será com Futuros."
    )

data_minima_permitida = datetime.date(2025, 1, 1)
data_maxima_permitida = datetime.date(2026, 9, 30)

with col_data_i:
    data_inicio = st.date_input(
        "Data Inicial", 
        min_value=data_minima_permitida, 
        max_value=data_maxima_permitida, 
        format="DD/MM/YYYY", 
        key="data_inicio_key_input_gas", 
        on_change=sincronizar_mes_pelas_datas,
        help="A partir de 01/01/2025. Se não modificar as datas ou o mês, será calculado a partir do dia seguinte ao atual."
    )
with col_data_f:
    data_fim = st.date_input(
        "Data Final", 
        min_value=data_minima_permitida, 
        max_value=data_maxima_permitida, 
        format="DD/MM/YYYY", 
        key="data_fim_key_input_gas", 
        on_change=sincronizar_mes_pelas_datas,
        help="De Data Inicial a 30/09/2026. Se não modificar as datas ou o mês, será calculado até um mês após a data inicial."
    )

# A lógica de cálculo dos dias lê os valores diretamente dos widgets
dias_default_calculado = (data_fim - data_inicio).days + 1 if data_fim >= data_inicio else 0

with col_dias_calc:
    gfx.exibir_metrica_personalizada("Dias (pelas datas)", f"{dias_default_calculado} dias")

with col_dias_man:
    # O valor por defeito do input manual é sempre o calculado a partir das datas
    dias_manual_input_val = st.number_input(
        "Nº Dias (manual)", 
        min_value=1, 
        value=dias_default_calculado, 
        step=1, 
        key="dias_manual_input_key_gas",
        help="Pode alterar os dias de forma manual, mas dê preferência às datas ou mês, para ter dados mais fidedignos nos tarifários indexados."
    )

# Decisão final sobre o número de dias a usar
if pd.isna(dias_manual_input_val) or dias_manual_input_val <= 0 or dias_manual_input_val == dias_default_calculado:
    dias = dias_default_calculado
else:
    dias = int(dias_manual_input_val)

st.write(f"Dias considerados: **{dias} dias**")

# Input MIBGAS
# Calcular o default ANTES de desenhar o widget
media_mibgas_calculada = 0.0
if not mibgas_df.empty:
    media_mibgas_calculada = calc.calcular_media_mibgas_datas(mibgas_df, data_inicio, data_fim)

# Se o cálculo falhar, usar o default das Constantes
if media_mibgas_calculada == 0.0:
     mibgas_default_const = calc.obter_constante("MIBGAS_Default", CONSTANTES)
     if mibgas_default_const != 0.0:
         media_mibgas_calculada = mibgas_default_const
     else:
         media_mibgas_calculada = 30.0 # Fallback final

# Guardar o default calculado para a callback do MIBGAS
st.session_state['mibgas_default_calculado'] = media_mibgas_calculada

# O valor default do input é a média que acabámos de calcular
mibgas_input_mwh = st.number_input(
    "Preço MIBGAS (€/MWh)",
    min_value=0.0,
    value=media_mibgas_calculada, # Default é o valor calculado
    step=1.0,
    format="%.2f",
    key="mibgas_input_mwh_manual", # Key para o input manual
    on_change=atualizar_url_mibgas,
    help=f"Valor médio esperado do MIBGAS. O valor pré-preenchido ({media_mibgas_calculada:.2f} €/MWh) é a média calculada para o período selecionado."
)

# Adicionar Alertas
if mibgas_input_mwh != media_mibgas_calculada:
    # Usar as funções de alerta (se existirem em graficos.py) ou st.info/st.warning
    diferenca_mibgas_perc = (mibgas_input_mwh / media_mibgas_calculada - 1) * 100 if media_mibgas_calculada != 0 else 0
    
    if mibgas_input_mwh < media_mibgas_calculada:
        st.warning(
            f"⚠️ **Aviso:** O valor MIBGAS inserido ({mibgas_input_mwh:.2f} €) é **{abs(diferenca_mibgas_perc):.1f}% inferior** à média calculada ({media_mibgas_calculada:.2f} €) para o período selecionado."
        )
    else:
        st.info(
            f"ℹ️ **Nota:** O valor MIBGAS inserido ({mibgas_input_mwh:.2f} €) é **{diferenca_mibgas_perc:.1f}% superior** à média calculada ({media_mibgas_calculada:.2f} €) para o período selecionado."
        )

# --- INÍCIO DO BLOCO DO GRÁFICO MIBGAS ---
with st.expander("📊 Ver Gráfico de Evolução dos Preços Médios Diários MIBGAS no Período"):
    
    # 1. Obter a data que separa os preços Spot dos Futuros (da aba 'Info')
    data_split_mibgas = None
    if not info_tab.empty:
        linha_data_split = info_tab[info_tab['Descricao'] == 'Ultima Data MIBGAS SPOT']
        if not linha_data_split.empty:
            valor_data = linha_data_split['Data'].iloc[0]
            if pd.notna(valor_data):
                try:
                    data_split_mibgas = pd.to_datetime(valor_data).date()
                except Exception:
                    st.warning("Não foi possível ler a data de referência dos valores MIBGAS SPOT.")

    # 2. Preparar e gerar o gráfico apenas se tivermos a data de split
    if data_split_mibgas:
        with st.spinner("A gerar gráfico MIBGAS..."):
            # Chamar a nova função de preparação de dados
            dados_grafico = gfx.preparar_dados_grafico_mibgas(
                df_mibgas=mibgas_df,
                data_inicio=data_inicio,
                data_fim=data_fim,
                data_split_spot_futuros=data_split_mibgas
            )
            
            # 3. Gerar e exibir o HTML do gráfico
            if dados_grafico:
                html_grafico_mibgas = gfx.gerar_grafico_highcharts_multi_serie(
                    chart_id=dados_grafico['id'],
                    chart_data=dados_grafico
                )
                st.components.v1.html(html_grafico_mibgas, height=320)
            else:
                st.info("Não existem dados MIBGAS disponíveis para o período selecionado para gerar o gráfico.")
    else:
        st.warning("A data de referência para os valores MIBGAS não está definida na aba 'Info' do ficheiro Excel.")
# --- FIM DO BLOCO DO GRÁFICO MIBGAS ---

# --- BLOCO DE INPUTS DE GÁS ---
st.markdown("##### Defina o seu perfil de consumo")

col_esc, col_mun, col_cur = st.columns(3)

with col_esc:
    # Escalão é selecionado (Default E1) e chama o callback
    escalao_map = {
        "Escalão 1 (Consumo até 220 m³/ano)": 1,
        "Escalão 2 (Consumo 221 a 500 m³/ano)": 2,
        "Escalão 3 (Consumo 501 a 1.000 m³/ano)": 3,
        "Escalão 4 (Consumo 1.001 a 10.000 m³/ano)": 4
    }
    escalao_selecionado_str = st.selectbox(
        "Selecione o seu Escalão de Consumo",
        options=list(escalao_map.keys()),
        index=0, # Default para Escalão 1
        key="sel_escalao_gas_key",
        on_change=atualizar_consumo_default_gas,
        help="Indique o seu **Escalão de Consumo** (pode encontrá-lo na sua fatura)"
    )
    escalao_num = escalao_map[escalao_selecionado_str]

with col_mun:
    # --- Definir default para o 10º item ---
    lista_municipios = sorted(tos_municipios['Município'].dropna().unique(), key=locale.strxfrm)
    
    # Definir o 10º item (índice 9) como padrão
    default_index_municipio = 9 
    # Safety check: se a lista tiver menos de 10 itens, usa o primeiro (índice 0)
    if len(lista_municipios) <= default_index_municipio: 
        default_index_municipio = 0
        
    municipio_selecionado = st.selectbox(
        "⚠️ Selecione o seu Município",
        options=lista_municipios,
        key="sel_municipio_tos",
        on_change=atualizar_url_municipio,
        help="O **Município** é essencial para calcular corretamente a Taxa de Ocupação do Subsolo (TOS)."
    )

# --- Mostrar o CUR correspondente ---
with col_cur:
    cur_selecionado_nome = ""
    linha_tos_cur = pd.DataFrame() # Inicializar como DataFrame vazio
    
    if municipio_selecionado:
        linha_tos_cur = tos_municipios[tos_municipios['Município'] == municipio_selecionado]
        if not linha_tos_cur.empty and 'CUR' in linha_tos_cur.columns:
            cur_selecionado_nome = linha_tos_cur.iloc[0].get('CUR', 'N/D')
    
    gfx.exibir_metrica_personalizada("Comercializador de Último Recurso (CUR)", cur_selecionado_nome if cur_selecionado_nome else "N/D")

# --- Inputs de Consumo (Com lógica de pré-preenchimento) ---
input_mode = st.radio(
    "Como prefere inserir o consumo?",
    ["Consumo (kWh)", "Consumo (m³)"],
    horizontal=True, index=0, key="gas_input_mode",
    on_change=atualizar_url_consumo_gas,
    help="**Consumo (kWh):** O valor final de energia que aparece na sua fatura. **Consumo (m³):** O volume de gás consumido, que também encontra na fatura. Terá de indicar o **Fator de Conversão (PCS)**, que converte m³ para kWh"
)

consumo_kwh = 0

# Verificamos se o callback já correu ou se é a primeira execução
if 'gas_kwh_input_key' not in st.session_state:
    atualizar_consumo_default_gas() # Chamar na primeira execução para definir o default (135)

if input_mode == "Consumo (kWh)":
    consumo_kwh = st.number_input(
        "Consumo (kWh)", 
        min_value=0, 
        step=10, 
        key="gas_kwh_input_key", # Key ligada ao callback
        on_change=atualizar_url_consumo_gas
    )
else:
    col_m3_1, col_m3_2 = st.columns(2)
    with col_m3_1:
        consumo_m3 = st.number_input("Consumo (m³)", min_value=0, value=12, step=1, key="gas_m3_input_key", on_change=atualizar_url_consumo_gas)
    with col_m3_2:
        fator_pcs = st.number_input("Fator de Conversão (PCS)", min_value=9.0, max_value=13.0, value=11.25, step=0.1, key="gas_pcs_input_key", on_change=atualizar_url_consumo_gas, help="Este valor (PCS) converte m³ para kWh e deve estar na sua fatura.")
    
    consumo_kwh = consumo_m3 * fator_pcs
    gfx.exibir_metrica_personalizada("Consumo (kWh) Calculado", f"{consumo_kwh:.0f} kWh")

# --- Calcular consumo pro-rata para o período ---
consumo_kwh_periodo_final = 0.0
if dias > 0 and consumo_kwh > 0:
    consumo_kwh_periodo_final = consumo_kwh / dias
    gfx.exibir_info_personalizada(f"O consumo médio diario para {dias} dias é de {consumo_kwh_periodo_final:.2f} kWh.")
elif dias <= 0:
    st.error("O número de dias da simulação não pode ser zero.")
    st.stop()
else:
    st.warning("Consumo é zero. Os cálculos de custo de energia e taxas variáveis serão zero.")


# --- Obter valores TOS ---
tos_fixo_dia_selecionado = 0.0
tos_variavel_kwh_selecionado = 0.0
if municipio_selecionado != "Outro":
    linha_tos = tos_municipios[tos_municipios['Município'] == municipio_selecionado]
    if not linha_tos.empty:
        try:
            if 'TOS_Fixo_Dia' in linha_tos.columns:
                 tos_fixo_dia_selecionado = float(linha_tos.iloc[0]['TOS_Fixo_Dia'])
            else:
                 tos_fixo_dia_selecionado = float(linha_tos.iloc[0]['TOS_Fixo_Dia (€/dia)']) # Fallback
                 
            if 'TOS_Variavel_kWh' in linha_tos.columns:
                tos_variavel_kwh_selecionado = float(linha_tos.iloc[0]['TOS_Variavel_kWh'])
            else:
                tos_variavel_kwh_selecionado = float(linha_tos.iloc[0]['TOS_Variavel_kWh (€/kWh)']) # Fallback
                
        except Exception as e:
            st.error(f"Erro ao ler valores TOS para {municipio_selecionado}. Verifique nomes das colunas na aba TOS. {e}")
            tos_fixo_dia_selecionado = 0.0
            tos_variavel_kwh_selecionado = 0.0


# --- Opções Adicionais ---
with st.expander("➕ Opções Adicionais de Simulação (Gás Natural)"):
    
    # --- Widget ISP ---
    isp_gas_default = calc.obter_constante('ISP_Gas_eur_kwh', CONSTANTES)
    
    isp_gas_manual_input = st.number_input(
        "ISP Gás (€/kWh)",
        min_value=0.0,
        value=isp_gas_default,
        step=0.0001,
        format="%.6f", 
        key="gas_isp_manual_input",
        on_change=atualizar_url_opcoes_adicionais_gas,
        help=f"Imposto Especial de Consumo (ISP). Default ({isp_gas_default})."
    )
    # TS Condicional
    if escalao_num in [1, 2]:
        st.markdown(r"##### Benefícios e Condições Especiais (para Escalões 1 e 2)") # Título condicional
        tarifa_social_gas = st.checkbox("Aplicar Tarifa Social de Gás Natural?", key="chk_ts_gas_v2", on_change=atualizar_url_opcoes_adicionais_gas, help="A Tarifa Social de Gás Natural aplica-se apenas aos escalões 1 e 2.")
    else:
        tarifa_social_gas = False
        if 'chk_ts_gas_v2' in st.session_state: 
            st.session_state.chk_ts_gas_v2 = False

    # Outros descontos
    st.markdown("##### Parcerias e Descontos Específicos")
    col_op1, col_op2 = st.columns(2)
    with col_op1:
        acp_gas = st.checkbox("Incluir quota ACP", key="chk_acp_gas", value=True, on_change=atualizar_url_opcoes_adicionais_gas, help="Inclui o valor da quota do ACP (4,80 €/mês) no valor do tarifário da parceria GE/ACP.",)
    with col_op2:
        desconto_continente_gas = st.checkbox("Desconto Continente", key="chk_cont_gas", value=True, on_change=atualizar_url_opcoes_adicionais_gas, help="Comparar o custo total incluindo o desconto do valor do cupão Continente no tarifário Galp&Continente.")


# --- "O Meu Tarifário" de Gás ---
help_O_Meu_Tarifario_Gas = """
Para preencher os valores de acordo com o seu tarifário, ou com outro qualquer que queira comparar.

**Atenção às notas sobre as TAR.**
    """
meu_tarifario_gas_ativo = st.checkbox("**Comparar com O Meu Tarifário de Gás Natural?**", key="chk_meu_tarifario_gas_ativo", on_change=atualizar_url_meu_tarifario_gas, help=help_O_Meu_Tarifario_Gas)

if meu_tarifario_gas_ativo:
    with st.container(border=True):

        st.subheader("🧾 O Meu Tarifário (Gás Natural)")
        
        col_meu1, col_meu2 = st.columns(2)
        with col_meu1:
            st.number_input("Preço Energia (€/kWh)", min_value=0.0, step=0.001, format="%g", key="meu_termo_energia_gas", on_change=atualizar_url_meu_tarifario_gas)
        with col_meu2:
            st.number_input("Preço Termo Fixo (€/dia)", min_value=0.0, step=0.001, format="%g", key="meu_termo_fixo_gas", on_change=atualizar_url_meu_tarifario_gas)

        col_meu_flag1, col_meu_flag2 = st.columns(2)
        with col_meu_flag1:
            st.checkbox("TAR incluída no Preço Energia?", value=True, key="meu_gas_tar_energia_incluida", on_change=atualizar_url_meu_tarifario_gas)
        with col_meu_flag2:
            st.checkbox("TAR incluída no Preço Termo Fixo?", value=True, key="meu_gas_tar_fixo_incluida", on_change=atualizar_url_meu_tarifario_gas)
            
        st.markdown("###### Descontos e Acréscimos (Meu Tarifário)")
        col_desc1, col_desc2, col_desc3, col_desc4 = st.columns(4)
        with col_desc1:
            st.number_input("Desconto Energia (%)", min_value=0.0, max_value=100.0, step=0.1, key="meu_gas_desconto_energia_perc", on_change=atualizar_url_meu_tarifario_gas)
        with col_desc2:
            st.number_input("Desconto Termo Fixo (%)", min_value=0.0, max_value=100.0, step=0.1, key="meu_gas_desconto_fixo_perc", on_change=atualizar_url_meu_tarifario_gas)
        with col_desc3:
            st.number_input("Desconto Fatura (€)", min_value=0.0, step=0.01, format="%.2f", key="meu_gas_desconto_fatura_eur", on_change=atualizar_url_meu_tarifario_gas)
        with col_desc4:
            st.number_input("Acréscimo Fatura (€)", min_value=0.0, step=0.01, format="%.2f", key="meu_gas_acrescimo_fatura_eur", on_change=atualizar_url_meu_tarifario_gas)

# --- BLOCO DO TARIFÁRIO PERSONALIZADO ---
help_Personalizado_Gas = """
Crie tarifário personalizado para comparar com os seus consumos. Ideal para comparar outro tarifário extra. Não permite descontos e acréscimos que existem em 'O Meu Tarifário'.

**Atenção às notas sobre as TAR.**
    """
personalizado_gas_ativo = st.checkbox(
    "**Comparar outro Tarifário Personalizado? (simplificado)**",
    key="chk_pers_gas_ativo", on_change=atualizar_url_tarifario_personalizado_gas,
    help=help_Personalizado_Gas
)

if personalizado_gas_ativo:
    with st.container(border=True):
        st.subheader("Tarifário Personalizado (Gás Natural)")
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            # O widget atualiza o session_state e o cálculo principal irá lê-lo.
            st.number_input("Preço Energia (€/kWh)", key="pers_gas_energia", min_value=0.0, step=0.001, format="%g", on_change=atualizar_url_tarifario_personalizado_gas)
        with col_s2:
            st.number_input("Preço Termo Fixo (€/dia)", key="pers_gas_fixo", min_value=0.0, step=0.001, format="%g", on_change=atualizar_url_tarifario_personalizado_gas)

        col_opt1, col_opt2 = st.columns(2)
        with col_opt1:
            st.checkbox("TAR incluída no Preço Energia?", value=True, key="pers_gas_tar_energia", on_change=atualizar_url_tarifario_personalizado_gas)
        with col_opt2:
            st.checkbox("TAR incluída no Preço Termo Fixo?", value=True, key="pers_gas_tar_potencia", on_change=atualizar_url_tarifario_personalizado_gas)

# --- FIM DO BLOCO PERSONALIZADO ---

# --- Helper Function ---
def get_filter_options_for_multiselect_gas(dataframe_gas, column_name):
    """Obtém opções únicas de filtro a partir da aba mestre de gás."""
    if column_name not in dataframe_gas.columns:
        return []
    options = dataframe_gas[column_name].dropna().unique()
    return sorted([opt for opt in options if opt and str(opt).lower() != 'pessoal'])

st.markdown("---") # Separador
st.subheader("🔍 Filtros da Tabela de Resultados")

# --- UI DE FILTROS ---
col_titulo_resultados, col_btn_limpar = st.columns([3,1]) # Ajustar a proporção conforme necessário

# --- Botão Limpar Filtros ---
with col_btn_limpar:
    st.write("") # Truque para alinhar verticalmente o botão (adiciona espaço no topo)
    st.write("")
    if st.button("🧹 Remover Todos os Filtros", key="btn_remover_filtros_gas", use_container_width=True, help="Remover Todos os Filtros"):
        # Keys de SelectBox (reset para índice 0)
        st.session_state.filter_segmento_gas_idx = 0  # Default "Residencial"
        st.session_state.filter_faturacao_gas_idx = 0 # Default "Todas"
        st.session_state.filter_pagamento_gas_idx = 0 # Default "Todos"
        # Key de MultiSelect (reset para lista vazia)
        st.session_state.filter_tipos_gas_multi = []
        st.rerun()

# Colocamos os 4 filtros e o botão de limpar, todos juntos.
filt_col1, filt_col2, filt_col3, filt_col4 = st.columns(4)

# --- Filtro de Segmento ---
with filt_col1:
    # 1. Definir a lista de opções
    opcoes_filtro_segmento_user = ["Residencial", "Empresarial", "Ambos"] 
    
    # 2. Encontrar o índice da opção default ("Residencial")
    default_index_segmento = opcoes_filtro_segmento_user.index("Residencial") # Isto será 0
    
    selected_segmento_user = st.selectbox(
        "Segmento", 
        opcoes_filtro_segmento_user, 
        index=st.session_state.get("filter_segmento_gas_idx", default_index_segmento), # Usa o índice de "Residencial"
        key="filter_segmento_gas_selectbox",
        help="Escolha o segmento para a simulação."
    )
    st.session_state.filter_segmento_gas_idx = opcoes_filtro_segmento_user.index(selected_segmento_user)

# --- Filtro de Tipo ---
with filt_col2:
    tipos_options_ms = get_filter_options_for_multiselect_gas(tarifas_gas_master, 'tipo') 
    
    help_text_formatado_gas = """
    Deixe em branco para mostrar todos os tipos.
    * **Fixo**: Preço de energia constante.
    * **Indexado**: Preço da energia baseado na média do MIBGAS para o período.
    """
    selected_tipos = st.multiselect("Tipo(s) de Tarifário", tipos_options_ms,
                                  default=st.session_state.get("filter_tipos_gas_multi", []),
                                  key="filter_tipos_gas_multi",
                                  help=help_text_formatado_gas)

# --- Filtro de Faturação ---
with filt_col3:
    opcoes_faturacao_user = ["Todas", "Fatura eletrónica", "Fatura em papel"]
    selected_faturacao_user = st.selectbox(
        "Faturação",
        opcoes_faturacao_user,
        index=st.session_state.get("filter_faturacao_gas_idx", 0), # Default "Todas"
        key="filter_faturacao_gas_selectbox",
        help="Escolha o tipo de faturação pretendido."
    )
    st.session_state.filter_faturacao_gas_idx = opcoes_faturacao_user.index(selected_faturacao_user)

# --- Filtro de Pagamento ---
with filt_col4:
    opcoes_pagamento_user = ["Todos", "Débito Direto", "Multibanco", "Numerário/Payshop/CTT"]
    selected_pagamento_user = st.selectbox(
        "Pagamento",
        opcoes_pagamento_user,
        index=st.session_state.get("filter_pagamento_gas_idx", 0), # Default "Todos"
        key="filter_pagamento_gas_selectbox",
        help="Escolha o método de pagamento."
    )
    st.session_state.filter_pagamento_gas_idx = opcoes_pagamento_user.index(selected_pagamento_user)
# --- FIM: UI DE FILTRO ---


# --- INÍCIO: LÓGICA DE FILTRAGEM PANDAS (PRÉ-CÁLCULO) ---

# --- Obter o Link do CUR para o Município selecionado ---
link_cur_municipio = ""
if not linha_tos_cur.empty and 'site_adesao' in linha_tos_cur.columns:
    link_cur_municipio = linha_tos_cur.iloc[0].get('site_adesao', '-')
    if pd.isna(link_cur_municipio):
        link_cur_municipio = "-"

with st.spinner("A filtrar e preparar dados..."):
    # 1. Limpeza de Tipos de Dados
    try:
        tarifas_gas_master['Escalão'] = pd.to_numeric(tarifas_gas_master['escalao'], errors='coerce').fillna(0).astype(int)
        tarifas_gas_master['Disponibilidade'] = tarifas_gas_master['disponibilidade'].str.strip().str.lower()
        cols_texto_filtro = ['segmento', 'tipo', 'faturacao', 'pagamento']
        for col in cols_texto_filtro:
            if col in tarifas_gas_master.columns:
                tarifas_gas_master[col] = tarifas_gas_master[col].astype(str).str.strip()
            
    except Exception as e:
        st.error(f"Erro ao tentar limpar os tipos de dados das tarifas de gás: {e}.")
        st.stop()

    # 2. Filtro PRIMÁRIO (Standalone e Escalão)
    df_gas_processar = tarifas_gas_master[
        (tarifas_gas_master['escalao'] == escalao_num) &
        ( (tarifas_gas_master['disponibilidade'] == 'g_so') | (tarifas_gas_master['disponibilidade'] == 'ambos') )
    ].copy()

    if df_gas_processar.empty and not meu_tarifario_gas_ativo:
         st.error(f"Não foram encontrados tarifários de gás standalone ('g_so' ou 'ambos') para o Escalão {escalao_num} na sua base de dados (antes dos filtros de tabela). Verifique os dados no Excel.")
         st.stop()

    # 3. Aplicar Filtros da UI
    df_a_filtrar = df_gas_processar.copy()

    # 3.1. Lógica para o filtro de Segmento
    if selected_segmento_user != "Ambos":
        segmentos_para_filtrar = []
        if selected_segmento_user == "Residencial":
            segmentos_para_filtrar.extend(["Doméstico", "Doméstico e Não Doméstico"]) 
        elif selected_segmento_user == "Empresarial":
            segmentos_para_filtrar.extend(["Não Doméstico", "Doméstico e Não Doméstico"])
        
        if 'segmento' in df_a_filtrar.columns:
            df_a_filtrar = df_a_filtrar[df_a_filtrar['segmento'].isin(segmentos_para_filtrar)]
        else:
            st.warning("Filtro 'Segmento' não aplicado: Coluna 'segmento' não encontrada na aba 'Tarifas_Gas_Master'.")

    # 3.2. Lógica para o filtro de Tipo (Multiselect)
    if selected_tipos: # Aplicar apenas se a lista NÃO estiver vazia
        if 'tipo' in df_a_filtrar.columns:
             df_a_filtrar = df_a_filtrar[df_a_filtrar['tipo'].isin(selected_tipos)]
        else:
             st.warning("Filtro 'Tipo' não aplicado: Coluna 'tipo' (minúscula) não encontrada.")

    # 3.3. Lógica para o filtro de Faturação
    if selected_faturacao_user != "Todas":
        faturacao_para_filtrar = []
        if selected_faturacao_user == "Fatura eletrónica":
            faturacao_para_filtrar.extend(["Fatura eletrónica", "Fatura eletrónica, Fatura em papel"])
        elif selected_faturacao_user == "Fatura em papel":
             faturacao_para_filtrar.extend(["Fatura eletrónica, Fatura em papel", "Fatura em papel"]) 
        
        if 'faturacao' in df_a_filtrar.columns:
            df_a_filtrar = df_a_filtrar[df_a_filtrar['faturacao'].isin(faturacao_para_filtrar)]
        else:
            st.warning("Filtro 'Faturação' não aplicado: Coluna 'faturacao' não encontrada.")

    # 3.4. Lógica para o filtro de Pagamento
    if selected_pagamento_user != "Todos":
        pagamento_para_filtrar = []
        if selected_pagamento_user == "Débito Direto":
            pagamento_para_filtrar.extend(["Débito Direto", "Débito Direto, Multibanco", "Débito Direto, Multibanco, Numerário/Payshop/CTT"])
        elif selected_pagamento_user == "Multibanco":
            pagamento_para_filtrar.extend(["Multibanco", "Débito Direto, Multibanco", "Débito Direto, Multibanco, Numerário/Payshop/CTT"])
        elif selected_pagamento_user == "Numerário/Payshop/CTT":
            pagamento_para_filtrar.extend(["Numerário/Payshop/CTT", "Débito Direto, Multibanco, Numerário/Payshop/CTT"])
        
        if 'pagamento' in df_a_filtrar.columns:
            df_a_filtrar = df_a_filtrar[df_a_filtrar['pagamento'].isin(pagamento_para_filtrar)]
        else:
            st.warning("Filtro 'Pagamento' não aplicado: Coluna 'pagamento' não encontrada.")

# --- FIM DA LÓGICA DE FILTRAGEM PANDAS ---

# --- INÍCIO DO CÁLCULO (Sobre o DF filtrado: df_a_filtrar) ---
with st.spinner(f"A calcular custos para os {len(df_a_filtrar)} tarifários de gás filtrados..."):
    
    resultados_list_gas = []
    
    # --- Definir se é um mês de faturação completo (para taxas fixas mensais) ---
    is_billing_month = 28 <= dias <= 31

    # Iterar e Calcular sobre o DataFrame JÁ FILTRADO
    for _, linha_tarifa in df_a_filtrar.iterrows():
        resultado_calculo = calc.calcular_custo_gas_completo(
            linha_tarifa,
            consumo_kwh, 
            dias,                      
            escalao_num,
            tarifa_social_gas,
            CONSTANTES,
            tos_fixo_dia_selecionado,
            tos_variavel_kwh_selecionado,
            mibgas_input_mwh,
            isp_gas_manual_input,
            acp_gas,
            desconto_continente_gas,
            VALOR_QUOTA_ACP_MENSAL
        )
        
        if resultado_calculo:
            # --- LÓGICA DE LINK DINÂMICO ---
            nome_tarifa_atual = str(resultado_calculo.get('NomeParaExibir', '')).lower()
            
            if "tarifa regulada" in nome_tarifa_atual and link_cur_municipio:
                # Se for tarifa regulada, usa o link do CUR do município
                resultado_calculo['LinkAdesao'] = link_cur_municipio
            else:
                # Caso contrário, usa o link (se existir) da aba Tarifas_Gas_Master
                resultado_calculo['LinkAdesao'] = linha_tarifa.get('site_adesao', '-')

            resultado_calculo['info_notas'] = linha_tarifa.get('notas', '')
            resultados_list_gas.append(resultado_calculo)

    # Calcular "O Meu Tarifário" (é calculado SEPARADAMENTE)
    if meu_tarifario_gas_ativo:
        resultado_meu_gas = calc.calcular_custo_meu_tarifario_gas(
            st.session_state,
            consumo_kwh,
            dias,
            escalao_num,
            tarifa_social_gas,
            CONSTANTES,
            tos_fixo_dia_selecionado,
            tos_variavel_kwh_selecionado,
            isp_gas_manual_input
        )
        if resultado_meu_gas:
            resultados_list_gas.append(resultado_meu_gas) # Adicionado à lista de resultados

    # --- CÁLCULO DO TARIFÁRIO PERSONALIZADO ---
    if personalizado_gas_ativo: # 'personalizado_gas_ativo' é o estado do checkbox
        
        # Criar o dicionário de inputs diretamente a partir das keys dos widgets (definidas no bloco UI)
        inputs_personalizado_gas = {
            'pers_gas_energia': st.session_state.get('pers_gas_energia', 0.0),
            'pers_gas_fixo': st.session_state.get('pers_gas_fixo', 0.0),
            'pers_gas_tar_energia': st.session_state.get('pers_gas_tar_energia', True),
            'pers_gas_tar_fixo': st.session_state.get('pers_gas_tar_potencia', True)
        }

        # Só calcular se houver algum preço definido
        if (float(inputs_personalizado_gas.get('pers_gas_energia', 0.0) or 0.0) > 0 or float(inputs_personalizado_gas.get('pers_gas_fixo', 0.0) or 0.0) > 0):
            
            resultado_personalizado_gas = calc.calcular_custo_personalizado_gas(
                inputs_personalizado_gas,
                consumo_kwh,
                dias,
                escalao_num,
                tarifa_social_gas,
                CONSTANTES,
                tos_fixo_dia_selecionado,
                tos_variavel_kwh_selecionado,
                isp_gas_manual_input
            )
            if resultado_personalizado_gas:
                resultados_list_gas.append(resultado_personalizado_gas)


    if not resultados_list_gas:
        st.warning("Nenhum tarifário corresponde aos filtros selecionados ou nenhum custo pôde ser calculado. Por favor, ajuste os filtros ou clique em 'Limpar'.")
        st.stop()
    
    st.markdown("---")

    # --- CONSTRUIR RESUMO DA SIMULAÇÃO ---
    cor_texto_resumo = "#333333" 
    resumo_html_parts = [
        f"<div style='background-color: #f9f9f9; border: 1px solid #ddd; padding: 15px; border-radius: 6px; margin-bottom: 25px; color: {cor_texto_resumo};'>"
    ]
    resumo_html_parts.append(f"<h5 style='margin-top:0; color: {cor_texto_resumo};'>Resumo da Simulação (Gás Natural):</h5>")
    resumo_html_parts.append("<ul style='list-style-type: none; padding-left: 0;'>")

    # Linha de Filtros
    linha_filtros = (
        f"<b>Segmento:</b> {selected_segmento_user} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Faturação:</b> {selected_faturacao_user} &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"<b>Pagamento:</b> {selected_pagamento_user}"
    )
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'>{linha_filtros}</li>")
    
    # Linha Escalão e Município
    linha_escalao_municipio = f"<b>{escalao_selecionado_str}</b> | <b>Município:</b> {municipio_selecionado}"
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'>{linha_escalao_municipio}</li>")

    # Linha de Consumo
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Consumo Total: {consumo_kwh:.0f} kWh</b></li>")

    # Linha Datas e Dias
    usou_dias_manuais = False
    if pd.notna(dias_manual_input_val) and dias_manual_input_val > 0 and int(dias_manual_input_val) != dias_default_calculado:
        usou_dias_manuais = True
    
    if usou_dias_manuais:
        resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Período:</b> {dias} dias (definido manualmente)</li>")
    else:
        resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>Período:</b> De {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')} ({dias} dias)</li>")

    # Linha MIBGAS
    resumo_html_parts.append(f"<li style='margin-bottom: 5px;'><b>MIBGAS:</b> {mibgas_input_mwh:.2f} €/MWh</li>")

    # Linha Tarifa Social
    if tarifa_social_gas:
        resumo_html_parts.append(f"<li style='margin-bottom: 5px; color: red;'><b>Benefício Aplicado:</b> Tarifa Social (Gás)</li>")

    resumo_html_parts.append("</ul></div>")
    html_resumo_final = "".join(resumo_html_parts)
    st.markdown(html_resumo_final, unsafe_allow_html=True)

    # --- TÍTULO DOS RESULTADOS ---
    # --- Processamento final e exibição da tabela de resultados ---
    st.subheader("💰 Tiago Felícia - Tarifários de Gás Natural")

    vista_simplificada = st.checkbox(
        "📱 Ativar vista simplificada (ideal em ecrãs menores)",
        value=True,
        key="chk_vista_simplificada_gas"
    )

    st.write("**Total** com todos os componentes, taxas e impostos e **valores unitários** de **Energia e Fixo** sem IVA.")
    st.write("**O nome do tarifário tem link para mais informações/adesão sobre o mesmo.**")

    st.markdown("➡️ [**Exportar Tabela para Excel**](#exportar-excel-detalhada-gas)")

    df_resultados_gas_final = pd.DataFrame(resultados_list_gas)
    
        # --- BLOCO PARA EXIBIR POUPANÇA ---
    try:
        # Inicializar/resetar variáveis do session_state (com keys específicas para gás)
        st.session_state.poupanca_excel_texto_gas = ""
        st.session_state.poupanca_excel_cor_gas = "000000"  # Preto
        st.session_state.poupanca_excel_negrito_gas = False

        if meu_tarifario_gas_ativo and not df_resultados_gas_final.empty:
            meu_tarifario_linha = df_resultados_gas_final[df_resultados_gas_final['NomeParaExibir'].str.contains("O Meu Tarifário", case=False, na=False)]

            if not meu_tarifario_linha.empty:
                custo_meu_tarifario = meu_tarifario_linha['Total Período (€)'].iloc[0]
                nome_meu_tarifario_ui = meu_tarifario_linha['NomeParaExibir'].iloc[0]

                if pd.notna(custo_meu_tarifario):
                    outros_tarifarios_ui_df = df_resultados_gas_final[
                        (df_resultados_gas_final['tipo'] != 'Pessoal')
                    ]
                    
                    nome_coluna_total = 'Total Período (€)'
                    
                    custos_outros_validos_ui = outros_tarifarios_ui_df[nome_coluna_total].dropna()

                    mensagem_poupanca_html_ui = "" 

                    if not custos_outros_validos_ui.empty:
                        custo_minimo_outros_ui = custos_outros_validos_ui.min()
                        linha_mais_barata_outros_ui = outros_tarifarios_ui_df.loc[outros_tarifarios_ui_df[nome_coluna_total] == custo_minimo_outros_ui].iloc[0]
                        nome_tarifario_mais_barato_outros_ui = linha_mais_barata_outros_ui['NomeParaExibir']

                        if custo_meu_tarifario > custo_minimo_outros_ui:
                            poupanca_abs_ui = custo_meu_tarifario - custo_minimo_outros_ui
                            poupanca_rel_ui = (poupanca_abs_ui / custo_meu_tarifario) * 100 if custo_meu_tarifario != 0 else 0
                            
                            mensagem_poupanca_html_ui = (
                                f"<span style='color:red; font-weight:bold;'>Poupança entre '{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f} €) e o mais económico da lista, "
                                f"'{nome_tarifario_mais_barato_outros_ui}' ({custo_minimo_outros_ui:.2f} €): </span>"
                                f"<span style='color:red; font-weight:bold;'>{poupanca_abs_ui:.2f} €</span> "
                                f"<span style='color:red; font-weight:bold;'>({poupanca_rel_ui:.2f} %).</span>"
                            )
                            # Guardar para Excel
                            st.session_state.poupanca_excel_texto_gas = (
                                f"Poupança entre '{nome_meu_tarifario_ui}' ({custo_meu_tarifario:.2f} €) e o mais económico da lista, "
                                f"'{nome_tarifario_mais_barato_outros_ui}' ({custo_minimo_outros_ui:.2f} €): "
                                f"{poupanca_abs_ui:.2f} € ({poupanca_rel_ui:.2f} %)."
                            )
                            st.session_state.poupanca_excel_cor_gas = "FF0000" # Vermelho
                            st.session_state.poupanca_excel_negrito_gas = True
                        
                        elif custo_meu_tarifario <= custo_minimo_outros_ui:
                            mensagem_poupanca_html_ui = f"<span style='color:green; font-weight:bold;'>Parabéns! O seu tarifário ('{nome_meu_tarifario_ui}' - {custo_meu_tarifario:.2f}€) já é o mais económico ou está entre os mais económicos da lista!</span>"
                            st.session_state.poupanca_excel_texto_gas = f"Parabéns! O seu tarifário ('{nome_meu_tarifario_ui}' - {custo_meu_tarifario:.2f}€) já é o mais económico ou está entre os mais económicos da lista!"
                            st.session_state.poupanca_excel_cor_gas = "008000" # Verde
                            st.session_state.poupanca_excel_negrito_gas = True
                    
                    if mensagem_poupanca_html_ui:
                        st.markdown(mensagem_poupanca_html_ui, unsafe_allow_html=True)

    except Exception as e_poupanca: 
        st.error(f"Erro ao processar a informação de poupança para UI (Gás): {e_poupanca}")
        st.session_state.poupanca_excel_texto_gas = "Erro ao calcular a informação de poupança."
    # --- FIM DO BLOCO DE POUPANÇA ---
    
    df_resultados_gas_final = df_resultados_gas_final.sort_values(by="Total Período (€)", ascending=True).reset_index(drop=True)

    # --- Lógica de Colunas Visíveis ---
    colunas_visiveis_presentes = []
    
    colunas_base_energia = ['Termo Energia (€/kWh)'] 
    coluna_fixo_gas = 'Termo Fixo (€/dia)'
    
    if vista_simplificada:
        colunas_base_visivel = ['NomeParaExibir', 'Total Período (€)']
        colunas_visiveis_presentes = colunas_base_visivel + colunas_base_energia
        if coluna_fixo_gas in df_resultados_gas_final.columns:
            colunas_visiveis_presentes.append(coluna_fixo_gas)
    else:
        colunas_base_visivel = ['NomeParaExibir', 'Total Período (€)']
        colunas_visiveis_presentes = colunas_base_visivel + colunas_base_energia
        if coluna_fixo_gas in df_resultados_gas_final.columns:
            colunas_visiveis_presentes.append(coluna_fixo_gas)
        colunas_visiveis_presentes.extend(['tipo', 'Comercializador', 'Segmento', 'Faturação', 'Pagamento'])

    colunas_visiveis_presentes = [col for col in colunas_visiveis_presentes if col in df_resultados_gas_final.columns]

    # --- Colunas Essenciais para JS (Tooltips e Estilos) ---
    colunas_dados_tooltip = [
        'tooltip_fixo_comerc_sem_tar', 'tooltip_fixo_tar_bruta', 'tooltip_fixo_ts_aplicada_flag', 'tooltip_fixo_ts_desconto_valor',
        'tooltip_energia_comerc_sem_tar', 'tooltip_energia_tar_bruta', 'tooltip_energia_ts_aplicada_flag', 'tooltip_energia_ts_desconto_valor',
        'tt_cte_energia_siva', 'tt_cte_fixo_siva', 'tt_cte_isp_siva', 'tt_cte_tos_fixo_siva', 'tt_cte_tos_var_siva',
        'tt_cte_total_siva', 'tt_cte_valor_iva_6_total', 'tt_cte_valor_iva_23_total',
        'tt_cte_subtotal_civa', 'tt_cte_desc_finais_valor', 'tt_cte_acres_finais_valor'
    ]
    colunas_essenciais_js = ['tipo', 'NomeParaExibir', 'LinkAdesao', 'info_notas'] 
    colunas_para_aggrid_final = list(dict.fromkeys(colunas_visiveis_presentes + colunas_essenciais_js + colunas_dados_tooltip))
    colunas_para_aggrid_final = [col for col in colunas_para_aggrid_final if col in df_resultados_gas_final.columns]

    df_aggrid_display = df_resultados_gas_final[colunas_para_aggrid_final].copy()

    gb = GridOptionsBuilder.from_dataframe(df_aggrid_display) 
    
    gb.configure_default_column(
        sortable=True,
        resizable=True,
        editable=False,
        wrapText=True,
        autoHeight=True,
        wrapHeaderText=True,
        autoHeaderHeight=True,
        headerClass="center-header"
    )

    # --- 1. DEFINIÇÕES JAVASCRIPT AVANÇADAS ---
    
    # Calcular Mín/Máx por Coluna
    cols_para_cor_gas = ['Total Período (€)', 'Termo Energia (€/kWh)', 'Termo Fixo (€/dia)']
    min_max_data_for_js = {}
    for col_name in cols_para_cor_gas:
        if col_name in df_aggrid_display:
            series = pd.to_numeric(df_aggrid_display[col_name], errors='coerce').dropna()
            if not series.empty:
                min_max_data_for_js[col_name] = {'min': series.min(), 'max': series.max()}
            else:
                min_max_data_for_js[col_name] = {'min': 0, 'max': 0}
    min_max_data_json_string = json.dumps(min_max_data_for_js)


    # Componente de Tooltip Personalizado
    custom_tooltip_component_js = JsCode("""
        class CustomTooltip { /* ...... */ 
            init(params) {
                this.eGui = document.createElement('div');
                this.eGui.innerHTML = params.value; 
                this.eGui.style.backgroundColor = 'white'; this.eGui.style.color = 'black';
                this.eGui.style.border = '1px solid #ccc'; this.eGui.style.padding = '10px';
                this.eGui.style.borderRadius = '6px'; this.eGui.style.boxShadow = '0 2px 5px rgba(0,0,0,0.15)';
                this.eGui.style.maxWidth = '400px'; this.eGui.style.fontSize = '1.1em';
                this.eGui.style.fontFamily = 'Arial, sans-serif'; this.eGui.style.whiteSpace = 'normal';
            }
            getGui() { return this.eGui; }
        }
    """)

    # Estilos de Célula para Nomes de Tarifário
    cor_fundo_indexado_gas_css = "#FFE699" 
    cor_texto_indexado_gas_css = "black"
    cor_fundo_fixo_gas_css = "#f0f0f0"     
    cor_texto_fixo_gas_css = "#333333"
    cor_fundo_personalizado_gas_css = "#92D050"
    cor_texto_personalizado_gas_css = "white"

    cell_style_nome_tarifario_js = JsCode(f"""
    function(params) {{
        let styleToApply = {{ textAlign: 'center', borderRadius: '6px', padding: '10px 10px' }};                                  
        if (params.data) {{
            const tipoTarifario = params.data.tipo; 
            const nomeTarifario = params.data.NomeParaExibir;

            if (tipoTarifario === 'Pessoal' && nomeTarifario && nomeTarifario.startsWith('O Meu Tarifário')) {{
                styleToApply.backgroundColor = 'red';
                styleToApply.color = 'white';
                styleToApply.fontWeight = 'bold';
            }} else if (tipoTarifario === 'Pessoal') {{ 
                /* Isto apanha o "Tarifário Personalizado" que também usa tipo 'Pessoal' */
                styleToApply.backgroundColor = '{cor_fundo_personalizado_gas_css}';
                styleToApply.color = '{cor_texto_personalizado_gas_css}';
                styleToApply.fontWeight = 'bold';
            }} else if (tipoTarifario === 'Indexado') {{
                styleToApply.backgroundColor = '{cor_fundo_indexado_gas_css}';
                styleToApply.color = '{cor_texto_indexado_gas_css}';                
            }} else if (tipoTarifario === 'Fixo') {{
                styleToApply.backgroundColor = '{cor_fundo_fixo_gas_css}';
                styleToApply.color = '{cor_texto_fixo_gas_css}';
            }}
            return styleToApply;
        }}
        return styleToApply; 
    }}
    """)

    # Tooltips JS para Termo Fixo e Energia
    tooltip_termo_fixo_gas_js = JsCode("""
    function(params) {
        if (!params.data) { return String(params.value); }
        const comercializador = parseFloat(params.data.tooltip_fixo_comerc_sem_tar || 0);
        const tarBruta = parseFloat(params.data.tooltip_fixo_tar_bruta || 0);
        const tsAplicada = params.data.tooltip_fixo_ts_aplicada_flag;
        const descontoTSValor = parseFloat(params.data.tooltip_fixo_ts_desconto_valor || 0);
        const formatPrice = (num) => (typeof num === 'number' && !isNaN(num)) ? num.toFixed(4) : 'N/A';
        let tooltipParts = ["<b>Decomposição Termo Fixo (s/IVA):</b>"];
        tooltipParts.push("Comercializador (s/TAR): " + formatPrice(comercializador) + " €/dia");
        tooltipParts.push("TAR (Tarifa Acesso Redes): " + formatPrice(tarBruta) + " €/dia");
        if (tsAplicada === true && descontoTSValor > 0) {
            tooltipParts.push("Desconto Tarifa Social: -" + formatPrice(descontoTSValor) + " €/dia");
        }
        tooltipParts.push("----------------------------------------------------");
        tooltipParts.push("<b>Custo Final : " + formatPrice(parseFloat(params.value)) + " €/dia</b>");
        return tooltipParts.join("<br>");
    }""")
    tooltip_termo_energia_gas_js = JsCode("""
    function(params) {
        if (!params.data) { return String(params.value); }
        const comercializador = parseFloat(params.data.tooltip_energia_comerc_sem_tar || 0);
        const tarBruta = parseFloat(params.data.tooltip_energia_tar_bruta || 0);
        const tsAplicada = params.data.tooltip_energia_ts_aplicada_flag;
        const descontoTSValor = parseFloat(params.data.tooltip_energia_ts_desconto_valor || 0);
        const formatPrice = (num) => (typeof num === 'number' && !isNaN(num)) ? num.toFixed(4) : 'N/A';
        let tooltipParts = ["<b>Decomposição Energia (s/IVA):</b>"];
        tooltipParts.push("Comercializador (s/TAR): " + formatPrice(comercializador) + " €/kWh");
        tooltipParts.push("TAR (Tarifa Acesso Redes): " + formatPrice(tarBruta) + " €/kWh");
        if (tsAplicada === true && descontoTSValor > 0) {
            tooltipParts.push("Desconto Tarifa Social: -" + formatPrice(descontoTSValor) + " €/kWh");
        }
        tooltipParts.push("----------------------------------------------------");
        tooltipParts.push("<b>Custo Final : " + formatPrice(parseFloat(params.value)) + " €/kWh</b>");
        return tooltipParts.join("<br>");
    }""")
    tooltip_custo_total_gas_js = JsCode("""
    function(params) {
        if (!params.data) { return String(params.value); }
        const formatCurrency = (num) => (typeof num === 'number' && !isNaN(num)) ? num.toFixed(2) : 'N/A';
        const nomeTarifario = params.data.NomeParaExibir || "Tarifário";
        let tooltipParts = [ "<i>" + nomeTarifario + "</i>", "<b>Decomposição Custo Total:</b>", "------------------------------------" ];
        const energia_siva = parseFloat(params.data.tt_cte_energia_siva || 0);
        const fixo_siva = parseFloat(params.data.tt_cte_fixo_siva || 0);
        const isp_siva = parseFloat(params.data.tt_cte_isp_siva || 0);
        const tos_fixo_siva = parseFloat(params.data.tt_cte_tos_fixo_siva || 0);
        const tos_var_siva = parseFloat(params.data.tt_cte_tos_var_siva || 0);
        const total_siva = parseFloat(params.data.tt_cte_total_siva || 0);
        const iva_6 = parseFloat(params.data.tt_cte_valor_iva_6_total || 0);
        const iva_23 = parseFloat(params.data.tt_cte_valor_iva_23_total || 0);
        const subtotal_civa = parseFloat(params.data.tt_cte_subtotal_civa || 0);
        const desc_finais = parseFloat(params.data.tt_cte_desc_finais_valor || 0);
        const acres_finais = parseFloat(params.data.tt_cte_acres_finais_valor || 0);
        tooltipParts.push("Total Energia s/IVA: " + formatCurrency(energia_siva) + " €");
        tooltipParts.push("Total Termo Fixo s/IVA: " + formatCurrency(fixo_siva) + " €");
        if (isp_siva !== 0) { tooltipParts.push("ISP s/IVA: " + formatCurrency(isp_siva) + " €"); }
        if (tos_fixo_siva !== 0 || tos_var_siva !== 0) { tooltipParts.push("Taxa Ocup. Subsolo (TOS): " + formatCurrency(tos_fixo_siva + tos_var_siva) + " €"); }
        tooltipParts.push("<b>Subtotal s/IVA: " + formatCurrency(total_siva) + " €</b>");
        tooltipParts.push("------------------------------------");
        if (iva_6 !== 0) { tooltipParts.push("Valor IVA (6%): " + formatCurrency(iva_6) + " €"); }
        if (iva_23 !== 0) { tooltipParts.push("Valor IVA (23%): " + formatCurrency(iva_23) + " €"); }
        tooltipParts.push("<b>Subtotal c/IVA: " + formatCurrency(subtotal_civa) + " €</b>");
        if (desc_finais !== 0 || acres_finais !== 0) {
            tooltipParts.push("------------------------------------");
            if (desc_finais !== 0) { tooltipParts.push("Outros Descontos: -" + formatCurrency(desc_finais) + " €"); }
            if (acres_finais !== 0) { tooltipParts.push("Outros Acréscimos: +" + formatCurrency(acres_finais) + " €"); }
            tooltipParts.push("------------------------------------");
        }
        tooltipParts.push("<b>Custo Total c/IVA: " + formatCurrency(parseFloat(params.value)) + " €</b>");
        return tooltipParts.join("<br>");
    }""")

    # Renderizador de Link/Notas
    link_tooltip_renderer_js = JsCode("""
        class LinkTooltipRenderer {
            init(params) {
                this.eGui = document.createElement('div');
                let displayText = params.value; 
                let url = params.data.LinkAdesao; 
                if (url && typeof url === 'string' && url.toLowerCase().startsWith('http')) {
                    this.eGui.innerHTML = `<a href="${url}" target="_blank" title="Aderir/Saber mais: ${url}" style="text-decoration: underline; color: inherit;">${displayText}</a>`;
                } else {
                    this.eGui.innerHTML = `<span title="${displayText}">${displayText}</span>`;
                }
            }
            getGui() { return this.eGui; }
        }
    """)
    
    # Getter de Tooltip de Notas
    tooltip_nome_tarifario_getter_js = JsCode("""
        function(params) {
            if (!params.data) { return params.value || ''; }
            const nomeExibir = params.data.NomeParaExibir || '';
            const notas = params.data.info_notas || ''; 
            let tooltipHtmlParts = [];
            if (nomeExibir) {
                tooltipHtmlParts.push("<strong>" + nomeExibir + "</strong>");
            }
            if (notas) {
                const notasHtml = notas.replace(/\\n/g, ' ').replace(/\n/g, ' ');
                tooltipHtmlParts.push("<small style='display: block; margin-top: 5px;'><i>" + notasHtml + "</i></small>");
            }
            if (tooltipHtmlParts.length > 0) {
                return tooltipHtmlParts.join('<br>');
            }
            return ''; 
        }
    """)

    # Gradiente de Cor Genérico
    cell_style_gradiente_custo_js = JsCode(f"""
    function(params) {{
        const colName = params.colDef.field;
        const value = parseFloat(params.value);
        const minMaxConfig = {min_max_data_json_string}; 
        
        let style = {{ textAlign: 'center', borderRadius: '6px', padding: '10px 10px' }}; 

        if (isNaN(value) || !minMaxConfig[colName]) {{ return style; }}
        const min_val = minMaxConfig[colName].min;
        const max_val = minMaxConfig[colName].max;
        if (max_val === min_val) {{ style.backgroundColor = 'lightgrey'; return style; }}
        
        const normalized_value = Math.max(0, Math.min(1, (value - min_val) / (max_val - min_val)));
        const cL={{r:99,g:190,b:123}},cM={{r:255,g:255,b:255}},cH={{r:248,g:105,b:107}}; 
        let r, g, b;
        if (normalized_value < 0.5) {{
            const t = normalized_value / 0.5; 
            r = Math.round(cL.r * (1 - t) + cM.r * t); g = Math.round(cL.g * (1 - t) + cM.g * t); b = Math.round(cL.b * (1 - t) + cM.b * t);
        }} else {{
            const t = (normalized_value - 0.5) / 0.5;
            r = Math.round(cM.r * (1 - t) + cH.r * t); g = Math.round(cM.g * (1 - t) + cH.g * t); b = Math.round(cM.b * (1 - t) + cH.b * t);
        }}
        style.backgroundColor = `rgb(${{r}},${{g}},${{b}})`;
        if ((r * 0.299 + g * 0.587 + b * 0.114) < 140) {{ 
            style.color = 'white';
        }} else {{
            style.color = 'black';
        }}
        return style;
    }}
    """)
    
    # --- Estilo de Linha para aplicar negrito ---
    get_row_style_js = JsCode("""
    function(params) {
        if (params.data && params.data.tipo === 'Pessoal') {
            // Aplica negrito a toda a linha se o 'tipo' for 'Pessoal'
            // (Isto apanha tanto 'O Meu Tarifário' como 'Tarifário Personalizado')
            return { 'fontWeight': 'bold' };
        }
        return null; // Sem estilo de linha para Fixo ou Indexado
    }
    """)


    # --- 2. CONFIGURAÇÃO DAS COLUNAS ---

    formatter_eur_5dec = JsCode("function(params) { if(params.value == null) return ''; return Number(params.value).toFixed(5); }") 
    formatter_eur_2dec = JsCode("function(params) { if(params.value == null) return ''; return '€ ' + Number(params.value).toFixed(2); }")

    is_visible_comerc = 'Comercializador' in colunas_visiveis_presentes 
    gb.configure_column("Comercializador", headerName="Comercializador", minWidth=150, flex=1.5, 
                        filter='agTextColumnFilter', cellStyle=cell_style_nome_tarifario_js,
                        hide=(not is_visible_comerc)) 
    
    # Coluna 2: Tarifário (Usa Link/Tooltip E Estilo de Cor)
    gb.configure_column("NomeParaExibir", headerName="Tarifário", minWidth=250, flex=2.5, filter='agTextColumnFilter', 
                        cellStyle=cell_style_nome_tarifario_js,
                        cellRenderer=link_tooltip_renderer_js,
                        tooltipValueGetter=tooltip_nome_tarifario_getter_js,
                        tooltipComponent=custom_tooltip_component_js)
    
    # Coluna 3: Total (Usa Gradiente (sem negrito) e Tooltip Total)
    gb.configure_column("Total Período (€)", headerName=f"Total ({dias} dias) (€)", type=["numericColumn"], 
                        valueFormatter=formatter_eur_2dec, 
                        cellStyle=cell_style_gradiente_custo_js, 
                        minWidth=130, flex=1,
                        tooltipValueGetter=tooltip_custo_total_gas_js,
                        tooltipComponent=custom_tooltip_component_js)

    # Coluna 4: Energia (Usa Gradiente (sem negrito) e Tooltip Energia)
    gb.configure_column("Termo Energia (€/kWh)", headerName="Energia (€/kWh)", type=["numericColumn"], 
                        valueFormatter=formatter_eur_5dec, minWidth=120, flex=1, 
                        cellStyle=cell_style_gradiente_custo_js, 
                        tooltipValueGetter=tooltip_termo_energia_gas_js,
                        tooltipComponent=custom_tooltip_component_js)
                        
    # Coluna 5: Fixo (Usa Gradiente (sem negrito) e Tooltip Fixo)
    gb.configure_column("Termo Fixo (€/dia)", headerName="Fixo (€/dia)", type=["numericColumn"], 
                        valueFormatter=formatter_eur_5dec, minWidth=120, flex=1, 
                        cellStyle=cell_style_gradiente_custo_js, 
                        tooltipValueGetter=tooltip_termo_fixo_gas_js,
                        tooltipComponent=custom_tooltip_component_js)

    # Colunas de Detalhe (com lógica hide)
    set_filter_params = { 'buttons': ['apply', 'reset'], 'excelMode': 'mac', 'suppressMiniFilter': False, }
    colunas_texto_detalhe = ['tipo', 'Segmento', 'Faturação', 'Pagamento']
    for col_name in colunas_texto_detalhe:
        if col_name in df_aggrid_display.columns:
            is_visible = col_name in colunas_visiveis_presentes
            header_name_display = 'Tipo' if col_name == 'tipo' else col_name 
            
            gb.configure_column(
                col_name, headerName=header_name_display, minWidth=120, flex=0.75,
                filter='agSetColumnFilter', filterParams=set_filter_params,
                cellStyle={'textAlign': 'center', 'backgroundColor': '#f0f0f0'},
                hide=(not is_visible) 
            )

    # Ocultar Colunas de Dados
    colunas_para_ocultar_final = [
        'LinkAdesao', 'info_notas', 
        'tooltip_fixo_comerc_sem_tar', 'tooltip_fixo_tar_bruta', 'tooltip_fixo_ts_aplicada_flag', 'tooltip_fixo_ts_desconto_valor',
        'tooltip_energia_comerc_sem_tar', 'tooltip_energia_tar_bruta', 'tooltip_energia_ts_aplicada_flag', 'tooltip_energia_ts_desconto_valor',
        'tt_cte_energia_siva', 'tt_cte_fixo_siva', 'tt_cte_isp_siva', 'tt_cte_tos_fixo_siva', 'tt_cte_tos_var_siva',
        'tt_cte_total_siva', 'tt_cte_valor_iva_6_total', 'tt_cte_valor_iva_23_total',
        'tt_cte_subtotal_civa', 'tt_cte_desc_finais_valor', 'tt_cte_acres_finais_valor'
    ]
    for col_ocultar in colunas_para_ocultar_final:
        if col_ocultar in df_aggrid_display.columns:
             gb.configure_column(col_ocultar, hide=True)

    gb.configure_grid_options(
        domLayout='autoHeight', # Para altura automática
        getRowStyle=get_row_style_js
    )

    gridOptions = gb.build()

    # --- CSS para centrar cabeçalhos ---
    custom_css = {
        ".ag-header-cell": {
            "display": "flex",
            "flex-direction": "column",
            "justify-content": "center !important",
            "align-items": "center !important",
            "text-align": "center !important"
        },
        ".ag-header-cell-label": {
            "justify-content": "center !important",
            "text-align": "center !important",
            "font-size": "14px !important",
            "font-weight": "bold !important"
        },
        ".ag-cell": {
            "font-size": "14px !important"
        },
        ".ag-center-cols-clip": {"justify-content": "center !important", "text-align": "center !important"}
    }

    # --- 3. RENDERIZAR O AGGRID ---
    grid_response = AgGrid(
        df_aggrid_display, 
        gridOptions=gridOptions,
        custom_css=custom_css,
        fit_columns_on_grid_load=True,
        theme='alpine',
        allow_unsafe_jscode=True,
        key="aggrid_gas_standalone_v16_bold_fix",
        enable_enterprise_modules=True,
        tooltipShowDelay=200, 
        tooltipMouseTrack=True,
        update_mode=GridUpdateMode.FILTERING_CHANGED | GridUpdateMode.SORTING_CHANGED
    )

    # --- BLOCO DE EXPORTAÇÃO EXCEL ---
    st.markdown("<a id='exportar-excel-detalhada-gas'></a>", unsafe_allow_html=True)
    st.markdown("---")
    with st.expander("📥 Exportar Tabela Detalhada para Excel"):

        opcoes_export_excel_gas = []
        default_cols_excel_gas = []

        if 'colunas_visiveis_presentes' in locals() and isinstance(colunas_visiveis_presentes, list):
            opcoes_export_excel_gas.extend(colunas_visiveis_presentes)
            default_cols_excel_gas = list(colunas_visiveis_presentes)

        colunas_tooltip_dados_para_export = colunas_para_ocultar_final 
        for col_tooltip in colunas_tooltip_dados_para_export:
            if col_tooltip in df_aggrid_display.columns and col_tooltip not in opcoes_export_excel_gas:
                opcoes_export_excel_gas.append(col_tooltip)

        colunas_para_exportar_excel_selecionadas = st.multiselect(
            "Selecione as colunas para exportar para Excel (Gás):",
            options=opcoes_export_excel_gas, 
            default=default_cols_excel_gas,
            key="cols_export_excel_selector_gas"
        )
        
        opcoes_limite_export = ["Todos"] + [f"Top {i}" for i in [10, 20, 30, 40, 50]]
        limite_export_selecionado = st.selectbox(
            "Número de tarifários a exportar (ordenados pelo 'Total (€)' atual da tabela):",
            options=opcoes_limite_export,
            index=0, 
            key="limite_tarifarios_export_excel_gas"
        )

        if st.button("Preparar Download do Ficheiro Excel (Gás)", key="btn_prep_excel_download_gas"):
            if not colunas_para_exportar_excel_selecionadas:
                st.warning("Por favor, selecione pelo menos uma coluna para exportar.")
            else:
                with st.spinner("A gerar ficheiro Excel de Gás..."):
                    
                    df_dados_filtrados_da_grid = pd.DataFrame()
                    if 'grid_response' in locals() and grid_response and grid_response['data'] is not None:
                        df_dados_filtrados_da_grid = pd.DataFrame(grid_response['data'])
                    else:
                        df_dados_filtrados_da_grid = df_aggrid_display.copy() 

                    if df_dados_filtrados_da_grid.empty and not df_aggrid_display.empty:
                        df_export_final = pd.DataFrame(columns=colunas_para_exportar_excel_selecionadas)
                    elif not df_dados_filtrados_da_grid.empty:
                        colunas_export_validas_no_filtrado = [col for col in colunas_para_exportar_excel_selecionadas if col in df_dados_filtrados_da_grid.columns]
                        if not colunas_export_validas_no_filtrado:
                            st.warning("Nenhuma das colunas selecionadas para exportação está presente nos dados filtrados.")
                            df_export_final = pd.DataFrame()
                        else:
                            df_export_final = df_dados_filtrados_da_grid[colunas_export_validas_no_filtrado].copy()
                    else: 
                        st.info("Tabela de Gás está vazia, nada para exportar.")
                        df_export_final = pd.DataFrame() 

                    if not df_export_final.empty and limite_export_selecionado != "Todos":
                        try:
                            num_a_exportar = int(limite_export_selecionado.split(" ")[1])
                            if len(df_export_final) > num_a_exportar:
                                df_export_final = df_export_final.head(num_a_exportar)
                        except Exception as e_limite_export:
                            st.warning(f"Não foi possível aplicar o limite de tarifários: {e_limite_export}")

                    if not df_export_final.empty:
                        nome_coluna_tarifario_excel = None
                        if 'NomeParaExibir' in df_export_final.columns:
                            df_export_final.rename(columns={'NomeParaExibir': 'Tarifário'}, inplace=True)
                            nome_coluna_tarifario_excel = 'Tarifário'
                        elif 'Tarifário' in df_export_final.columns:
                            nome_coluna_tarifario_excel = 'Tarifário'

                        tipos_reais_para_estilo = None
                        if 'tipo' in df_dados_filtrados_da_grid.columns:
                            try:
                                tipos_reais_para_estilo = df_dados_filtrados_da_grid.loc[df_export_final.index, 'tipo']
                            except KeyError:
                                tipos_reais_para_estilo = pd.Series(index=df_export_final.index, dtype=str)
                        else:
                            tipos_reais_para_estilo = pd.Series(index=df_export_final.index, dtype=str)

                        # Arredondar dados
                        for col in df_export_final.columns:
                            if col in ['Total Período (€)']:
                                df_export_final[col] = pd.to_numeric(df_export_final[col], errors='coerce').round(2)
                            elif col in ['Termo Energia (€/kWh)', 'Termo Fixo (€/dia)']:
                                 df_export_final[col] = pd.to_numeric(df_export_final[col], errors='coerce').round(5)
                        
                        styler_excel = df_export_final.style.apply(
                            lambda df: estilo_geral_dataframe_para_exportar(df, tipos_reais_para_estilo, min_max_data_for_js, nome_coluna_tarifario_excel), 
                            axis=None
                        )
                        
                        styler_excel = styler_excel.format(formatter="{:.2f}", subset=["Total Período (€)"], na_rep="-")
                        styler_excel = styler_excel.format(formatter="{:.5f}", subset=['Termo Energia (€/kWh)', 'Termo Fixo (€/dia)'], na_rep="-")
                        
                        styler_excel = styler_excel.set_table_styles([
                            {'selector': 'th', 'props': [('background-color', '#404040'), ('color', 'white'), ('font-weight', 'bold'), ('text-align', 'center'), ('border', '1px solid black'), ('padding', '5px')]},
                            {'selector': 'td', 'props': [('border', '1px solid #dddddd'), ('padding', '4px')]}
                        ]).hide(axis="index")

                        # Passar a flag 'personalizado_gas_ativo' para a função
                        output_excel_bytes = exportar_excel_completo(
                            df_export_final,
                            styler_excel,
                            html_resumo_final,
                            st.session_state.get('poupanca_excel_texto_gas', ""),
                            escalao_selecionado_str, 
                            meu_tarifario_gas_ativo,
                            personalizado_gas_ativo
                        )

                        timestamp_final_dl = int(time.time())
                        nome_ficheiro_final_dl = f"Tiago_Felicia_Gas_Natural_{timestamp_final_dl}.xlsx"
            
                        st.download_button(
                            label=f"📥 Descarregar Excel ({nome_ficheiro_final_dl})",
                            data=output_excel_bytes.getvalue(),
                            file_name=nome_ficheiro_final_dl,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"btn_dl_excel_gas_completo_{timestamp_final_dl}" 
                        )
                        st.success(f"{nome_ficheiro_final_dl} pronto para download!")
    
    # --- FIM DO BLOCO DE EXPORTAÇÃO EXCEL ---

    # --- INÍCIO: PÓDIO DA POUPANÇA ---
    st.subheader("🏆 O Seu Pódio da Poupança (Gás)")
    st.markdown("Estas são as 3 opções mais económicas para si, com base nos seus consumos atuais.")

    # Garantir que o DataFrame está ordenado e o índice está correto
    df_resultados_ordenado_gas = df_resultados_gas_final.sort_values(by="Total Período (€)").reset_index(drop=True)
    top3_gas = df_resultados_ordenado_gas.head(3)

    if len(top3_gas) >= 3:
        custo_referencia = None
        nome_referencia = ""
    
        if meu_tarifario_gas_ativo and not df_resultados_ordenado_gas[df_resultados_ordenado_gas['NomeParaExibir'].str.contains("O Meu Tarifário", na=False)].empty:
            meu_tar_resultado = df_resultados_ordenado_gas[df_resultados_ordenado_gas['NomeParaExibir'].str.contains("O Meu Tarifário", na=False)].iloc[0]
            if 'Total Período (€)' in meu_tar_resultado and pd.notna(meu_tar_resultado['Total Período (€)']):
                custo_referencia = meu_tar_resultado['Total Período (€)']
                nome_referencia = meu_tar_resultado['NomeParaExibir']

        if custo_referencia is None and not df_resultados_ordenado_gas.empty:
            # Fallback: Usar o mais caro se o Meu Tarifário não existir
            pior_tarifario = df_resultados_ordenado_gas.iloc[-1]
            custo_referencia = pior_tarifario['Total Período (€)']
            nome_referencia = pior_tarifario['NomeParaExibir']
    
        if custo_referencia is not None:
            st.caption(f"A comparação é feita em relação ao seu ponto de referência ('{nome_referencia}' ({custo_referencia:.2f} €)).")
    
            col2, col1, col3 = st.columns([1, 1.2, 1])

            def apresentar_item_podio_gas(coluna, dados_podio, emoji):
                with coluna:
                    st.markdown(f"<p style='text-align: center; font-size: 24px;'>{emoji}</p>", unsafe_allow_html=True)
                    with st.container(border=True):
                        st.markdown(f"<p style='text-align: center; font-weight: bold;'>{dados_podio['NomeParaExibir']}</p>", unsafe_allow_html=True)
                        st.metric("Custo Estimado", f"{dados_podio['Total Período (€)']:.2f} €")
                    
                        diferenca = dados_podio['Total Período (€)'] - custo_referencia
                        if diferenca < 0:
                            st.metric("Poupança", f"{abs(diferenca):.2f} €/mês", delta_color="off")
                        elif diferenca > 0:
                            st.metric("Custo Adicional", f"{diferenca:.2f} €/mês", delta=f"{diferenca:.2f} €", delta_color="inverse")
                        else:
                            st.metric("Custo", "Igual à referência", delta_color="off")

            apresentar_item_podio_gas(col1, top3_gas.iloc[0], "🥇 1º lugar")
            apresentar_item_podio_gas(col2, top3_gas.iloc[1], "🥈2º lugar")
            apresentar_item_podio_gas(col3, top3_gas.iloc[2], "🥉3º lugar")

    st.markdown("---")
    # --- FIM: PÓDIO DA POUPANÇA ---

# --- SECÇÃO DE PARTILHA POR URL
st.subheader("🔗 Partilhar Simulação")

# A lógica só é executada se existirem parâmetros no URL
if st.query_params:
    
    # --- URL para o da sua página de Gás Natural ---
    base_url = "https://tiagofelicia-gas.streamlit.app/" 

    query_string = "&".join([f"{k}={v}" for k, v in st.query_params.items()])
    shareable_link = f"{base_url}?{query_string}"

    # --- Componente HTML/JS para o campo de texto e botão de copiar ---
    html_componente_copiar = f"""
    <div style="display: flex; align-items: center; gap: 8px; font-family: sans-serif;">
        <input 
            type="text" 
            id="shareable-link-input-gas" 
            value="{shareable_link}" 
            readonly 
            style="width: 100%; padding: 8px; border-radius: 6px; border: 1px solid #ccc; font-size: 14px;"
        >
        <button 
            id="copy-button-gas" 
            onclick="copyLinkToClipboardGas()"
            style="
                padding: 8px 12px; 
                border-radius: 6px; 
                border: 1px solid #ccc;
                background-color: #f0f2f6; 
                cursor: pointer;
                font-size: 14px;
                white-space: nowrap;
            "
        >
            📋 Copiar Link
        </button>
    </div>

    <script>
    function copyLinkToClipboardGas() {{
        const linkInput = document.getElementById("shareable-link-input-gas");
        linkInput.select();
        linkInput.setSelectionRange(0, 99999); // Para telemóveis

        navigator.clipboard.writeText(linkInput.value).then(() => {{
            const copyButton = document.getElementById("copy-button-gas");
            copyButton.innerText = "Copiado!";
            setTimeout(() => {{
                copyButton.innerHTML = "&#128203; Copiar Link";
            }}, 2000);
        }}).catch(err => {{
            console.error('Falha ao copiar o link: ', err);
        }});
    }}
    </script>
    """
    st.components.v1.html(html_componente_copiar, height=55)

else:
    st.info("Altere um dos parâmetros da simulação (Escalão, Município, Consumo, etc.) para gerar um link de partilha.")


# Legenda das Colunas da Tabela Tarifários de Gás Natural
st.markdown("---")
st.subheader("📖 Legenda das Colunas da Tabela Tarifários de Gás Natural")
st.caption("""
* **Tarifário**: Nome identificativo do tarifário. Pode incluir notas sobre descontos de fatura específicos.
* **Tipo**: Indica se o tarifário é:
    * `Fixo`: Preços de energia e potência são constantes.
    * `Indexado`: Preço da energia baseado na média do MIBGÁS para o período selecionado.
    * `Pessoal`: O seu tarifário, conforme introduzido.
* **Comercializador**: Empresa que oferece o tarifário.
* **[...] (€/kWh)**: Custo unitário da energia, **sem IVA**.
    * Para "O Meu Tarifário", este valor já reflete quaisquer descontos percentuais de energia e o desconto da Tarifa Social que tenhas configurado.
    * Para os outros tarifários, é o preço base sem IVA, já considerando o desconto da Tarifa Social se ativa.
* **Termo Fixo (€/dia)**: Custo unitário diário do Termo Fixo **sem IVA**.
    * Para "O Meu Tarifário", este valor já reflete quaisquer descontos percentuais de potência e o desconto da Tarifa Social que tenhas configurado.
    * Para os outros tarifários, é o preço base sem IVA, já considerando o desconto da Tarifa Social se ativa.
* **Total (€)**: Valor do custo final estimado da fatura para o período simulado. Este custo inclui:
    * Custo da energia consumida (com IVA aplicado conforme as regras).
    * Custo do Termo Fixo (com IVA aplicado conforme as regras).
    * Taxas adicionais: ISP Gás e TOS.
    * Quaisquer descontos de fatura em euros (para "O Meu Tarifário" ou especificados nos tarifários).
""")

# Usar f-strings para construir o HTML da legenda
st.subheader("🎨 Legenda de Cores por Tipo de Tarifário")

# Definições de cores
cor_fundo_meu_tarifario_legenda = "red"
cor_texto_meu_tarifario_legenda = "white"
cor_fundo_tarifario_personalizado_legenda = "#92D050"
cor_texto_tarifario_personalizado_legenda = "white"
cor_fundo_indexado_css = "#FFE699"
cor_texto_indexado_media_css = "black"
cor_fundo_fixo_legenda = "#f0f0f0" # Cor ligeiramente cinza para Fixo
cor_texto_fixo_legenda = "#333333"
borda_fixo_legenda = "#CCCCCC"     # Borda para o quadrado branco ser visível

# Construir a string HTML da legenda passo a passo
html_items = ""

# Item: O Meu Tarifário
if meu_tarifario_gas_ativo:
    html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
        <div style="width: 18px; height: 18px; background-color: {cor_fundo_meu_tarifario_legenda}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
        <span style="background-color: {cor_fundo_meu_tarifario_legenda}; color: {cor_texto_meu_tarifario_legenda}; padding: 2px 6px; border-radius: 4px; font-weight: bold;">O Meu Tarifário</span>
        <span style="margin-left: 8px;">- Tarifário configurado pelo utilizador.</span>
    </div>"""

# Item: Tarifário Personalizado
if personalizado_gas_ativo:
    html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
        <div style="width: 18px; height: 18px; background-color: {cor_fundo_tarifario_personalizado_legenda}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
        <span style="background-color: {cor_fundo_tarifario_personalizado_legenda}; color: {cor_texto_tarifario_personalizado_legenda}; padding: 2px 6px; border-radius: 4px; font-weight: bold;">Tarifário Personalizado</span>
        <span style="margin-left: 8px;">- Tarifário configurado pelo utilizador.</span>
    </div>"""

# Item: Indexado
html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
    <div style="width: 18px; height: 18px; background-color: {cor_fundo_indexado_css}; border: 1px solid #ccc; border-radius: 4px; margin-right: 8px;"></div>
    <span style="background-color: {cor_fundo_indexado_css}; color: {cor_texto_indexado_media_css}; padding: 2px 6px; border-radius: 4px;">Indexado</span>
    <span style="margin-left: 8px;">- Preço de energia baseado na média MIBGÁS do período definido.</span>
</div>"""

# Item: Fixo
html_items += f"""<div style="display: flex; align-items: center; margin-bottom: 5px;">
    <div style="width: 18px; height: 18px; background-color: {cor_fundo_fixo_legenda}; border: 1px solid {borda_fixo_legenda}; border-radius: 4px; margin-right: 8px;"></div>
    <span style="background-color: {cor_fundo_fixo_legenda}; color: {cor_texto_fixo_legenda}; padding: 2px 6px; border-radius: 4px;">Tarifário Fixo</span>
    <span style="margin-left: 8px;">- Preços de energia constantes.</span>
</div>"""

# Criar o HTML final e renderizar
legenda_html_completa = f"""<div style="font-size: 14px;">{html_items}</div>"""
st.markdown(legenda_html_completa, unsafe_allow_html=True)

# --- DATAS DE REFERÊNCIA ---
st.markdown("---") # Adiciona um separador visual
st.subheader("📅 Data de Referência dos Valores de Mercado no simulador")

# Processar e exibir Data_Valores_MIBGAS_SPOT a partir da aba "Info"
data_mibgas_spot_formatada_str = "Não disponível"

linha_data_mibgas_spot = info_tab[info_tab['Descricao'] == 'Ultima Data MIBGAS SPOT']

if not linha_data_mibgas_spot.empty:
    data_valores_mibgas_spot = linha_data_mibgas_spot['Data'].iloc[0]
    if pd.notna(data_valores_mibgas_spot):
        data_valores_mibgas_spot_temp = None # Variável temporária para a data MIBGAS
        try:
            if isinstance(data_valores_mibgas_spot, (datetime.datetime, pd.Timestamp)):
                data_valores_mibgas_spot_temp = data_valores_mibgas_spot.date()
            else:
                timestamp_convertido_mibgas = pd.to_datetime(data_valores_mibgas_spot, errors='coerce')
                if pd.notna(timestamp_convertido_mibgas):
                    data_valores_mibgas_spot_temp = timestamp_convertido_mibgas.date()
            
            if data_valores_mibgas_spot_temp and isinstance(data_valores_mibgas_spot_temp, datetime.date):
                data_mibgas_spot_formatada_str = data_valores_mibgas_spot_temp.strftime('%d/%m/%Y')
            elif data_valores_mibgas_spot:
                data_mibgas_spot_formatada_str = f"Valor não reconhecido como data ({data_valores_mibgas_spot})"
        except Exception: # Captura outros erros de conversão
            if data_valores_mibgas_spot:
                data_mibgas_spot_formatada_str = f"Erro ao processar valor ({data_valores_mibgas_spot})"

st.markdown(f"**Valores MIBGAS (SPOT) atualizados em** {data_mibgas_spot_formatada_str}")
# --- FIM DA SECÇÃO ---

# --- INÍCIO DA SECÇÃO DE APOIO ---
st.markdown("---") # Adiciona um separador visual antes da secção de apoio
st.subheader("💖 Apoie este Projeto")

st.markdown(
    "Se quiser apoiar a manutenção do site e o desenvolvimento contínuo deste simulador, "
    "pode fazê-lo através de uma das seguintes formas:"
)

# Link para BuyMeACoffee
st.markdown(
    "☕ [**Compre-me um café em BuyMeACoffee**](https://buymeacoffee.com/tiagofelicia)"
)

st.markdown("ou através do botão PayPal:")

# Código HTML para o botão do PayPal
paypal_button_html = """
<div style="text-align: left; margin-top: 10px; margin-bottom: 15px;">
    <form action="https://www.paypal.com/donate" method="post" target="_blank" style="display: inline-block;">
    <input type="hidden" name="hosted_button_id" value="W6KZHVL53VFJC">
    <input type="image" src="https://www.paypalobjects.com/pt_PT/PT/i/btn/btn_donate_SM.gif" border="0" name="submit" title="PayPal - The safer, easier way to pay online!" alt="Faça donativos com o botão PayPal">
    <img alt="" border="0" src="https://www.paypal.com/pt_PT/i/scr/pixel.gif" width="1" height="1">
    </form>
</div>
"""
st.markdown(paypal_button_html, unsafe_allow_html=True)
# --- FIM DA SECÇÃO DE APOIO ---

st.markdown("---")
# Título para as redes sociais
st.subheader("Redes sociais, onde poderão seguir o projeto:")

# URLs das redes sociais
url_x = "https://x.com/tiagofelicia"
url_bluesky = "https://bsky.app/profile/tiagofelicia.bsky.social"
url_youtube = "https://youtube.com/@tiagofelicia"
url_facebook_perfil = "https://www.facebook.com/profile.php?id=61555007360529"


icon_url_x = "https://upload.wikimedia.org/wikipedia/commons/thumb/c/cc/X_icon.svg/120px-X_icon.svg.png?20250519203220"
icon_url_bluesky = "https://upload.wikimedia.org/wikipedia/commons/7/7a/Bluesky_Logo.svg"
icon_url_youtube = "https://upload.wikimedia.org/wikipedia/commons/thumb/f/fd/YouTube_full-color_icon_%282024%29.svg/120px-YouTube_full-color_icon_%282024%29.svg.png"
icon_url_facebook = "https://upload.wikimedia.org/wikipedia/commons/thumb/b/b9/2023_Facebook_icon.svg/120px-2023_Facebook_icon.svg.png"


svg_icon_style_dark_mode_friendly = "filter: invert(0.8) sepia(0) saturate(1) hue-rotate(0deg) brightness(1.5) contrast(0.8);"

col_social1, col_social2, col_social3, col_social4 = st.columns(4)

with col_social1:
    st.markdown(
        f"""
        <a href="{url_x}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_x}" width="40" alt="X" style="margin-bottom: 8px; object-fit: contain;">
            X
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social2:
    st.markdown(
        f"""
        <a href="{url_bluesky}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_bluesky}" width="40" alt="Bluesky" style="margin-bottom: 8px; object-fit: contain;">
            Bluesky
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social3:
    st.markdown(
        f"""
        <a href="{url_youtube}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_youtube}" width="40" alt="YouTube" style="margin-bottom: 8px; object-fit: contain;">
            YouTube
        </a>
        """,
        unsafe_allow_html=True
    )

with col_social4:
    st.markdown(
        f"""
        <a href="{url_facebook_perfil}" target="_blank" style="text-decoration: none; color: inherit; display: flex; flex-direction: column; align-items: center; text-align: center;">
            <img src="{icon_url_facebook}" width="40" alt="Facebook" style="margin-bottom: 8px; object-fit: contain;">
            Facebook
        </a>
        """,
        unsafe_allow_html=True
    )

st.markdown("<br>", unsafe_allow_html=True) # Adiciona um espaço vertical

# Texto de Copyright
ano_copyright = 2025
nome_autor = "Tiago Felícia"
texto_copyright_html = f"© {ano_copyright} Todos os direitos reservados | {nome_autor} | <a href='{url_facebook_perfil}' target='_blank' style='color: inherit;'>Facebook</a>"

st.markdown(
    f"<div style='text-align: center; font-size: 0.9em; color: grey;'>{texto_copyright_html}</div>",
    unsafe_allow_html=True
)
